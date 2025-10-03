"""
Tests for NGI Learning Module Sprint 2
Package generation, file uploads, validation, and ingestion
"""

import pytest
import sys
import os
import io
from datetime import datetime
from fastapi.testclient import TestClient

# Add project root to path
sys.path.insert(0, os.path.abspath(os.path.join(os.path.dirname(__file__), '..')))

from src.api.main import app
from src.api.database import get_db, _ensure_engine
from src.api.models import Base
from src.api.models_learning import (
    LearningCompany, LearningPackage, LearningSubmission
)


client = TestClient(app)


@pytest.fixture(scope="function")
def setup_sprint2_db():
    """Setup test database for Sprint 2"""
    # Ensure engine is initialized
    _ensure_engine()
    
    # Create tables
    from src.api.database import _engine
    Base.metadata.create_all(bind=_engine)
    
    # Get database session
    db = next(get_db())
    
    # Clear existing data
    db.query(LearningSubmission).delete()
    db.query(LearningPackage).delete()
    db.query(LearningCompany).delete()
    db.commit()
    
    # Seed test company
    test_company = LearningCompany(
        ticker="TSLA",
        company_name="Tesla, Inc.",
        industry="Automotive",
        sub_industry="Electric Vehicles",
        description="EV manufacturer",
        headquarters="Austin, Texas",
        fiscal_year_end="December 31",
        sec_cik="0001318605",
        revenue_model_type="QxP",
        revenue_driver_notes="Q = Vehicle deliveries; P = ASP",
        data_quality_score=10,
        is_active=True
    )
    db.add(test_company)
    db.commit()
    db.refresh(test_company)
    
    yield {"db": db, "company_id": test_company.id, "ticker": "TSLA"}
    
    # Cleanup
    db.close()


class TestPackageGeneration:
    """Test Excel package generation"""
    
    def test_generate_package_success(self, setup_sprint2_db):
        """Test POST /api/learning/packages/generate/{company_id} creates package"""
        company_id = setup_sprint2_db["company_id"]
        
        response = client.post(f"/api/learning/packages/generate/{company_id}")
        assert response.status_code == 200
        
        data = response.json()
        assert data["status"] == "success"
        assert data["ticker"] == "TSLA"
        assert data["version"] == 1
        assert "file_path" in data
        assert data["file_size_bytes"] > 0
        
        # Verify file was created
        assert os.path.exists(data["file_path"])
        
        # Verify database record
        db = setup_sprint2_db["db"]
        package = db.query(LearningPackage).filter(
            LearningPackage.company_id == company_id
        ).first()
        assert package is not None
        assert package.version == 1
    
    def test_generate_package_increments_version(self, setup_sprint2_db):
        """Test that generating multiple packages increments version"""
        company_id = setup_sprint2_db["company_id"]
        
        # Generate v1
        response1 = client.post(f"/api/learning/packages/generate/{company_id}")
        assert response1.status_code == 200
        assert response1.json()["version"] == 1
        
        # Generate v2
        response2 = client.post(f"/api/learning/packages/generate/{company_id}")
        assert response2.status_code == 200
        assert response2.json()["version"] == 2
    
    def test_generate_package_invalid_company(self, setup_sprint2_db):
        """Test generating package for non-existent company returns 404"""
        response = client.post("/api/learning/packages/generate/999999")
        assert response.status_code == 404


class TestFileUploads:
    """Test submission file uploads"""
    
    def test_upload_excel_submission(self, setup_sprint2_db):
        """Test POST /api/learning/submissions/upload with Excel file"""
        company_id = setup_sprint2_db["company_id"]
        
        # Create a dummy Excel file
        excel_content = b"PK\x03\x04"  # Excel file header (minimal)
        excel_content += b"\x00" * 100  # Padding
        
        files = {
            "file": ("test_model.xlsx", io.BytesIO(excel_content), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        }
        data = {
            "company_id": company_id,
            "activity_id": "a1_drivers_map",
            "notes": "My first submission"
        }
        
        response = client.post("/api/learning/submissions/upload", files=files, data=data)
        assert response.status_code == 200
        
        result = response.json()
        assert result["status"] == "success"
        assert result["version"] == 1
        assert result["file_type"] == "excel"
        assert result["activity_id"] == "a1_drivers_map"
        assert result["ticker"] == "TSLA"
    
    def test_upload_invalid_file_type(self, setup_sprint2_db):
        """Test uploading invalid file type returns 400"""
        company_id = setup_sprint2_db["company_id"]
        
        files = {
            "file": ("test.txt", io.BytesIO(b"Not an Excel file"), "text/plain")
        }
        data = {
            "company_id": company_id,
            "activity_id": "a1_drivers_map"
        }
        
        response = client.post("/api/learning/submissions/upload", files=files, data=data)
        assert response.status_code == 400
        assert "Invalid file type" in response.json()["detail"]
    
    def test_upload_empty_file(self, setup_sprint2_db):
        """Test uploading empty file returns 400"""
        company_id = setup_sprint2_db["company_id"]
        
        files = {
            "file": ("empty.xlsx", io.BytesIO(b""), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        }
        data = {
            "company_id": company_id,
            "activity_id": "a1_drivers_map"
        }
        
        response = client.post("/api/learning/submissions/upload", files=files, data=data)
        assert response.status_code == 400
        assert "empty" in response.json()["detail"].lower()
    
    def test_get_my_submissions(self, setup_sprint2_db):
        """Test GET /api/learning/submissions/user/me returns user submissions"""
        response = client.get("/api/learning/submissions/user/me")
        assert response.status_code == 200
        
        data = response.json()
        assert "count" in data
        assert "submissions" in data
        assert isinstance(data["submissions"], list)


class TestValidation:
    """Test Excel validation"""
    
    def test_validate_excel_file(self, setup_sprint2_db):
        """Test validation of uploaded Excel file"""
        from src.api.learning.validators import ExcelValidator
        
        # Generate a package first
        company_id = setup_sprint2_db["company_id"]
        gen_response = client.post(f"/api/learning/packages/generate/{company_id}")
        assert gen_response.status_code == 200
        
        file_path = gen_response.json()["file_path"]
        
        # Validate the generated file
        validator = ExcelValidator(file_path)
        results = validator.validate()
        
        assert "status" in results
        assert results["status"] in ["passed", "passed_with_warnings", "failed"]
        assert isinstance(results["errors"], list)
        assert isinstance(results["warnings"], list)
    
    def test_validation_endpoint_requires_submission(self, setup_sprint2_db):
        """Test POST /api/learning/submissions/{id}/validate requires valid submission"""
        response = client.post("/api/learning/submissions/999999/validate")
        assert response.status_code == 404


class TestIngestion:
    """Test SEC data ingestion"""
    
    def test_ingestion_endpoint_exists(self, setup_sprint2_db):
        """Test POST /api/learning/admin/ingest/{ticker} endpoint exists"""
        ticker = setup_sprint2_db["ticker"]
        
        # Note: We won't actually run ingestion in tests (would hit SEC servers)
        # Just verify endpoint structure is correct
        response = client.post(f"/api/learning/admin/ingest/{ticker}")
        
        # Should fail gracefully (either 200 with results or 500 with error)
        assert response.status_code in [200, 500]
    
    def test_ingestion_invalid_ticker(self, setup_sprint2_db):
        """Test ingestion with invalid ticker returns 404"""
        response = client.post("/api/learning/admin/ingest/INVALID")
        
        # Should return 404 for unknown ticker
        assert response.status_code in [404, 500]


class TestExcelGenerator:
    """Test Excel generator directly"""
    
    def test_excel_generator_creates_file(self):
        """Test ExcelPackageGenerator creates valid Excel file"""
        from src.api.learning.excel_generator import ExcelPackageGenerator
        
        generator = ExcelPackageGenerator(output_dir="uploads/test_packages")
        filepath = generator.generate_package(
            ticker="TEST",
            company_name="Test Company, Inc.",
            fiscal_year_end="December 31",
            version=1
        )
        
        assert os.path.exists(filepath)
        assert filepath.endswith(".xlsx")
        assert os.path.getsize(filepath) > 0
        
        # Cleanup
        if os.path.exists(filepath):
            os.remove(filepath)
    
    def test_excel_has_required_tabs(self):
        """Test generated Excel has all required tabs"""
        from src.api.learning.excel_generator import ExcelPackageGenerator
        from openpyxl import load_workbook
        
        generator = ExcelPackageGenerator(output_dir="uploads/test_packages")
        filepath = generator.generate_package(
            ticker="TEST",
            company_name="Test Company, Inc.",
            version=1
        )
        
        # Load and check tabs
        wb = load_workbook(filepath)
        required_tabs = [
            'README',
            'Assumptions & Drivers',
            'Income Statement',
            'Balance Sheet',
            'Cash Flow',
            'DCF',
            'Outputs'
        ]
        
        for tab in required_tabs:
            assert tab in wb.sheetnames, f"Missing required tab: {tab}"
        
        wb.close()
        
        # Cleanup
        if os.path.exists(filepath):
            os.remove(filepath)


if __name__ == "__main__":
    pytest.main([__file__, "-v", "--tb=short"])


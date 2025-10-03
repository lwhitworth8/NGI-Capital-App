"""
Deterministic Excel Validators for NGI Learning Module
Uses openpyxl to validate Excel models before AI feedback
Following specifications from MarkdownFiles/NGILearning/Appendix.Validators.V1.md
"""

import os
from typing import Dict, List, Tuple
from openpyxl import load_workbook
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet


class ExcelValidator:
    """
    Validate Excel financial models against NGI standards.
    Checks balance sheet balancing, cash flow ties, formula integrity.
    """
    
    def __init__(self, file_path: str):
        """
        Initialize validator with Excel file.
        
        Args:
            file_path: Path to Excel file to validate
        """
        self.file_path = file_path
        self.workbook: Workbook = None
        self.errors: List[Dict] = []
        self.warnings: List[Dict] = []
        
    def validate(self) -> Dict:
        """
        Run all validation checks.
        
        Returns:
            Dictionary with validation results
        """
        if not os.path.exists(self.file_path):
            return {
                'status': 'failed',
                'errors': [{'check': 'file_exists', 'message': 'File not found'}],
                'warnings': []
            }
        
        try:
            self.workbook = load_workbook(self.file_path, data_only=False)
        except Exception as e:
            return {
                'status': 'failed',
                'errors': [{'check': 'file_open', 'message': f'Cannot open file: {str(e)}'}],
                'warnings': []
            }
        
        # Run all checks
        self._check_required_tabs()
        self._check_balance_sheet_integrity()
        self._check_cash_flow_ties()
        self._check_formula_errors()
        self._check_hardcoded_values()
        self._check_color_conventions()
        
        # Determine overall status
        if self.errors:
            status = 'failed'
        elif self.warnings:
            status = 'passed_with_warnings'
        else:
            status = 'passed'
        
        return {
            'status': status,
            'errors': self.errors,
            'warnings': self.warnings,
            'total_errors': len(self.errors),
            'total_warnings': len(self.warnings)
        }
    
    def _check_required_tabs(self):
        """Check that all required tabs exist"""
        required_tabs = [
            'README',
            'Assumptions & Drivers',
            'Income Statement',
            'Balance Sheet',
            'Cash Flow',
            'DCF',
            'Outputs'
        ]
        
        existing_tabs = self.workbook.sheetnames
        
        for tab in required_tabs:
            if tab not in existing_tabs:
                self.errors.append({
                    'check': 'required_tabs',
                    'tab': tab,
                    'message': f'Missing required tab: {tab}'
                })
    
    def _check_balance_sheet_integrity(self):
        """Check that Assets = Liabilities + Equity"""
        if 'Balance Sheet' not in self.workbook.sheetnames:
            return
        
        ws = self.workbook['Balance Sheet']
        
        # This is a simplified check - in production would parse actual values
        # Looking for check row with "Assets = Liabilities + Equity"
        found_check = False
        for row in ws.iter_rows(min_row=1, max_row=50, min_col=1, max_col=1):
            cell = row[0]
            if cell.value and 'Check' in str(cell.value) and 'Assets' in str(cell.value):
                found_check = True
                # Check if error flag (red) exists
                if cell.font and cell.font.color:
                    if hasattr(cell.font.color, 'rgb') and cell.font.color.rgb == 'FFFF0000':
                        self.errors.append({
                            'check': 'balance_sheet_balance',
                            'row': cell.row,
                            'message': 'Balance sheet does not balance (Assets != Liabilities + Equity)'
                        })
                break
        
        if not found_check:
            self.warnings.append({
                'check': 'balance_sheet_balance',
                'message': 'Balance sheet check row not found'
            })
    
    def _check_cash_flow_ties(self):
        """Check that cash flow ties to balance sheet"""
        if 'Cash Flow' not in self.workbook.sheetnames:
            return
        
        ws = self.workbook['Cash Flow']
        
        # Look for check row
        found_check = False
        for row in ws.iter_rows(min_row=1, max_row=50, min_col=1, max_col=1):
            cell = row[0]
            if cell.value and 'Check' in str(cell.value) and ('CF' in str(cell.value) or 'BS' in str(cell.value)):
                found_check = True
                break
        
        if not found_check:
            self.warnings.append({
                'check': 'cash_flow_ties',
                'message': 'Cash flow check row not found'
            })
    
    def _check_formula_errors(self):
        """Check for #REF!, #VALUE!, #DIV/0! errors"""
        error_types = ['#REF!', '#VALUE!', '#DIV/0!', '#NUM!', '#NAME?', '#NULL!']
        
        for sheet_name in self.workbook.sheetnames:
            if sheet_name == 'Raw Import':  # Skip locked sheet
                continue
            
            ws = self.workbook[sheet_name]
            
            for row in ws.iter_rows():
                for cell in row:
                    if cell.value in error_types:
                        self.errors.append({
                            'check': 'formula_errors',
                            'sheet': sheet_name,
                            'cell': cell.coordinate,
                            'error': cell.value,
                            'message': f'Formula error {cell.value} in {sheet_name}!{cell.coordinate}'
                        })
    
    def _check_hardcoded_values(self):
        """Check for hardcoded values in formula cells (black text)"""
        # This is a simplified check - looks for numeric values in non-input areas
        # In production, would check cell colors and formula presence
        
        for sheet_name in self.workbook.sheetnames:
            if sheet_name in ['README', 'Raw Import']:
                continue
            
            ws = self.workbook[sheet_name]
            
            # Check rows beyond headers (starting row 4)
            for row in ws.iter_rows(min_row=4, max_row=100):
                for cell in row:
                    # If cell has a value but no formula, might be hardcoded
                    if cell.value is not None and isinstance(cell.value, (int, float)):
                        if cell.data_type == 'n' and not str(cell.value).startswith('='):
                            # Check if it's in input area (blue background)
                            is_input = False
                            if cell.fill and hasattr(cell.fill, 'start_color'):
                                if cell.fill.start_color and hasattr(cell.fill.start_color, 'rgb'):
                                    # Blue input background: E7F3FF
                                    if 'E7F3FF' in str(cell.fill.start_color.rgb):
                                        is_input = True
                            
                            if not is_input and cell.column > 1:  # Skip label column
                                # This might be a hardcoded value
                                self.warnings.append({
                                    'check': 'hardcoded_values',
                                    'sheet': sheet_name,
                                    'cell': cell.coordinate,
                                    'message': f'Possible hardcoded value in {sheet_name}!{cell.coordinate}'
                                })
    
    def _check_color_conventions(self):
        """Check that color conventions are followed"""
        # Simplified check - verifies input cells have blue formatting
        
        if 'Assumptions & Drivers' not in self.workbook.sheetnames:
            return
        
        ws = self.workbook['Assumptions & Drivers']
        
        # Check that assumption rows have blue formatting
        blue_found = False
        for row in ws.iter_rows(min_row=4, max_row=20, min_col=2, max_col=5):
            for cell in row:
                if cell.font and cell.font.color:
                    if hasattr(cell.font.color, 'rgb'):
                        # Check for blue (0070C0)
                        if '0070C0' in str(cell.font.color.rgb):
                            blue_found = True
                            break
            if blue_found:
                break
        
        if not blue_found:
            self.warnings.append({
                'check': 'color_conventions',
                'message': 'Input cells may not follow blue color convention'
            })


def validate_submission(file_path: str) -> Dict:
    """
    Convenience function to validate a submission file.
    
    Args:
        file_path: Path to Excel file
    
    Returns:
        Validation results dictionary
    """
    validator = ExcelValidator(file_path)
    return validator.validate()


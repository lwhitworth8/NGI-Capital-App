"""
Excel Validation System - Goldman Sachs Investment Banking Standards
==================================================================

This module provides comprehensive Excel validation for financial models,
ensuring compliance with investment banking standards and best practices.

Author: NGI Capital Learning Team
"""

import openpyxl
import xlsxwriter
from typing import Dict, List, Tuple, Any, Optional
from dataclasses import dataclass
from enum import Enum
import re
import logging

logger = logging.getLogger(__name__)

class ValidationSeverity(Enum):
    """Validation severity levels"""
    ERROR = "error"
    WARNING = "warning"
    INFO = "info"

@dataclass
class ValidationResult:
    """Result of a validation check"""
    severity: ValidationSeverity
    message: str
    cell: Optional[str] = None
    sheet: Optional[str] = None
    suggestion: Optional[str] = None

class ExcelValidator:
    """Excel validation system for investment banking standards"""
    
    def __init__(self):
        self.results: List[ValidationResult] = []
        self.workbook = None
        self.worksheet = None
        
    def validate_workbook(self, file_path: str) -> List[ValidationResult]:
        """
        Validate an Excel workbook against investment banking standards
        
        Args:
            file_path: Path to the Excel file
            
        Returns:
            List of validation results
        """
        self.results = []
        
        try:
            self.workbook = openpyxl.load_workbook(file_path)
            self._validate_workbook_structure()
            self._validate_worksheets()
            self._validate_formulas()
            self._validate_formatting()
            self._validate_data_integrity()
            self._validate_financial_model_structure()
            
        except Exception as e:
            self.results.append(ValidationResult(
                severity=ValidationSeverity.ERROR,
                message=f"Failed to load workbook: {str(e)}"
            ))
        
        return self.results
    
    def _validate_workbook_structure(self):
        """Validate basic workbook structure"""
        # Check for required sheets
        required_sheets = ["Summary", "Assumptions", "Income Statement", "Balance Sheet", "Cash Flow"]
        existing_sheets = self.workbook.sheetnames
        
        for sheet in required_sheets:
            if sheet not in existing_sheets:
                self.results.append(ValidationResult(
                    severity=ValidationSeverity.WARNING,
                    message=f"Missing recommended sheet: {sheet}",
                    suggestion="Consider adding this sheet for better organization"
                ))
    
    def _validate_worksheets(self):
        """Validate individual worksheets"""
        for sheet_name in self.workbook.sheetnames:
            worksheet = self.workbook[sheet_name]
            self._validate_worksheet(worksheet, sheet_name)
    
    def _validate_worksheet(self, worksheet, sheet_name: str):
        """Validate a single worksheet"""
        # Check for proper headers
        self._validate_headers(worksheet, sheet_name)
        
        # Check for hardcoded values in formulas
        self._validate_hardcoded_values(worksheet, sheet_name)
        
        # Check for circular references
        self._validate_circular_references(worksheet, sheet_name)
        
        # Check for proper number formatting
        self._validate_number_formatting(worksheet, sheet_name)
        
        # Check for data validation
        self._validate_data_validation(worksheet, sheet_name)
    
    def _validate_headers(self, worksheet, sheet_name: str):
        """Validate worksheet headers"""
        # Check for merged cells in headers
        merged_ranges = list(worksheet.merged_cells.ranges)
        if merged_ranges:
            self.results.append(ValidationResult(
                severity=ValidationSeverity.WARNING,
                message="Worksheet contains merged cells",
                sheet=sheet_name,
                suggestion="Avoid merged cells for better data integrity"
            ))
        
        # Check for proper header formatting
        for row in range(1, min(6, worksheet.max_row + 1)):
            for col in range(1, min(10, worksheet.max_column + 1)):
                cell = worksheet.cell(row=row, column=col)
                if cell.value and isinstance(cell.value, str):
                    if cell.value.strip() and not cell.font.bold:
                        self.results.append(ValidationResult(
                            severity=ValidationSeverity.INFO,
                            message=f"Header cell {cell.coordinate} should be bold",
                            cell=cell.coordinate,
                            sheet=sheet_name
                        ))
    
    def _validate_hardcoded_values(self, worksheet, sheet_name: str):
        """Validate for hardcoded values in formulas"""
        for row in worksheet.iter_rows():
            for cell in row:
                if cell.data_type == 'f' and cell.value:  # Formula cell
                    formula = str(cell.value)
                    # Check for hardcoded numbers in formulas
                    hardcoded_pattern = r'[=+\-*/]\s*\d+\.?\d*'
                    if re.search(hardcoded_pattern, formula):
                        self.results.append(ValidationResult(
                            severity=ValidationSeverity.WARNING,
                            message=f"Hardcoded value in formula at {cell.coordinate}",
                            cell=cell.coordinate,
                            sheet=sheet_name,
                            suggestion="Consider using named ranges or cell references"
                        ))
    
    def _validate_circular_references(self, worksheet, sheet_name: str):
        """Validate for circular references"""
        # This is a simplified check - in practice, you'd need more sophisticated logic
        for row in worksheet.iter_rows():
            for cell in row:
                if cell.data_type == 'f' and cell.value:
                    formula = str(cell.value)
                    cell_ref = cell.coordinate
                    if cell_ref in formula:
                        self.results.append(ValidationResult(
                            severity=ValidationSeverity.ERROR,
                            message=f"Potential circular reference at {cell.coordinate}",
                            cell=cell.coordinate,
                            sheet=sheet_name
                        ))
    
    def _validate_number_formatting(self, worksheet, sheet_name: str):
        """Validate number formatting"""
        for row in worksheet.iter_rows():
            for cell in row:
                if cell.data_type == 'n' and cell.value is not None:
                    # Check for proper number formatting
                    if cell.number_format == 'General':
                        self.results.append(ValidationResult(
                            severity=ValidationSeverity.INFO,
                            message=f"Number cell {cell.coordinate} should have specific formatting",
                            cell=cell.coordinate,
                            sheet=sheet_name,
                            suggestion="Use appropriate number format (e.g., #,##0 for integers)"
                        ))
    
    def _validate_data_validation(self, worksheet, sheet_name: str):
        """Validate data validation rules"""
        # Check for data validation on key cells
        for row in worksheet.iter_rows():
            for cell in row:
                if cell.data_validation:
                    # Data validation exists - this is good
                    continue
                elif cell.value and isinstance(cell.value, (int, float)):
                    # Numeric cell without validation
                    self.results.append(ValidationResult(
                        severity=ValidationSeverity.INFO,
                        message=f"Consider adding data validation to {cell.coordinate}",
                        cell=cell.coordinate,
                        sheet=sheet_name
                    ))
    
    def _validate_formulas(self):
        """Validate formulas across the workbook"""
        for sheet_name in self.workbook.sheetnames:
            worksheet = self.workbook[sheet_name]
            self._validate_formula_consistency(worksheet, sheet_name)
            self._validate_formula_complexity(worksheet, sheet_name)
    
    def _validate_formula_consistency(self, worksheet, sheet_name: str):
        """Validate formula consistency"""
        # Check for consistent formulas in ranges
        for row in range(1, worksheet.max_row + 1):
            for col in range(1, worksheet.max_column + 1):
                cell = worksheet.cell(row=row, column=col)
                if cell.data_type == 'f' and cell.value:
                    # Check if formula is consistent with adjacent cells
                    self._check_formula_consistency(worksheet, cell, sheet_name)
    
    def _check_formula_consistency(self, worksheet, cell, sheet_name: str):
        """Check if formula is consistent with adjacent cells"""
        # This is a simplified check - in practice, you'd need more sophisticated logic
        formula = str(cell.value)
        if 'SUM(' in formula or 'AVERAGE(' in formula:
            # Check if adjacent cells have similar formulas
            adjacent_cells = [
                worksheet.cell(row=cell.row-1, column=cell.column),
                worksheet.cell(row=cell.row+1, column=cell.column),
                worksheet.cell(row=cell.row, column=cell.column-1),
                worksheet.cell(row=cell.row, column=cell.column+1)
            ]
            
            similar_formulas = 0
            for adj_cell in adjacent_cells:
                if adj_cell.data_type == 'f' and adj_cell.value:
                    if str(adj_cell.value) == formula:
                        similar_formulas += 1
            
            if similar_formulas == 0:
                self.results.append(ValidationResult(
                    severity=ValidationSeverity.INFO,
                    message=f"Formula at {cell.coordinate} may not be consistent with adjacent cells",
                    cell=cell.coordinate,
                    sheet=sheet_name
                ))
    
    def _validate_formula_complexity(self, worksheet, sheet_name: str):
        """Validate formula complexity"""
        for row in worksheet.iter_rows():
            for cell in row:
                if cell.data_type == 'f' and cell.value:
                    formula = str(cell.value)
                    # Check for overly complex formulas
                    if len(formula) > 200:
                        self.results.append(ValidationResult(
                            severity=ValidationSeverity.WARNING,
                            message=f"Complex formula at {cell.coordinate} may be hard to maintain",
                            cell=cell.coordinate,
                            sheet=sheet_name,
                            suggestion="Consider breaking down into smaller, more manageable formulas"
                        ))
    
    def _validate_formatting(self):
        """Validate formatting standards"""
        for sheet_name in self.workbook.sheetnames:
            worksheet = self.workbook[sheet_name]
            self._validate_cell_formatting(worksheet, sheet_name)
            self._validate_color_coding(worksheet, sheet_name)
    
    def _validate_cell_formatting(self, worksheet, sheet_name: str):
        """Validate cell formatting"""
        for row in worksheet.iter_rows():
            for cell in row:
                if cell.value is not None:
                    # Check for proper alignment
                    if cell.alignment.horizontal is None:
                        self.results.append(ValidationResult(
                            severity=ValidationSeverity.INFO,
                            message=f"Cell {cell.coordinate} should have explicit alignment",
                            cell=cell.coordinate,
                            sheet=sheet_name
                        ))
                    
                    # Check for proper borders
                    if not cell.border.left.style:
                        self.results.append(ValidationResult(
                            severity=ValidationSeverity.INFO,
                            message=f"Cell {cell.coordinate} should have borders for better readability",
                            cell=cell.coordinate,
                            sheet=sheet_name
                        ))
    
    def _validate_color_coding(self, worksheet, sheet_name: str):
        """Validate color coding standards"""
        # Check for consistent color coding
        colors_used = set()
        for row in worksheet.iter_rows():
            for cell in row:
                if cell.fill.start_color.rgb:
                    colors_used.add(cell.fill.start_color.rgb)
        
        if len(colors_used) > 10:
            self.results.append(ValidationResult(
                severity=ValidationSeverity.WARNING,
                message=f"Too many colors used in {sheet_name}",
                sheet=sheet_name,
                suggestion="Use a consistent color scheme for better readability"
            ))
    
    def _validate_data_integrity(self):
        """Validate data integrity"""
        for sheet_name in self.workbook.sheetnames:
            worksheet = self.workbook[sheet_name]
            self._validate_data_types(worksheet, sheet_name)
            self._validate_data_ranges(worksheet, sheet_name)
    
    def _validate_data_types(self, worksheet, sheet_name: str):
        """Validate data types"""
        for row in worksheet.iter_rows():
            for cell in row:
                if cell.value is not None:
                    # Check for mixed data types in columns
                    if isinstance(cell.value, str) and cell.value.replace('.', '').replace('-', '').isdigit():
                        self.results.append(ValidationResult(
                            severity=ValidationSeverity.WARNING,
                            message=f"Text value that looks like a number at {cell.coordinate}",
                            cell=cell.coordinate,
                            sheet=sheet_name,
                            suggestion="Convert to number format"
                        ))
    
    def _validate_data_ranges(self, worksheet, sheet_name: str):
        """Validate data ranges"""
        for row in worksheet.iter_rows():
            for cell in row:
                if cell.data_type == 'n' and cell.value is not None:
                    # Check for reasonable ranges
                    if abs(cell.value) > 1e12:
                        self.results.append(ValidationResult(
                            severity=ValidationSeverity.WARNING,
                            message=f"Very large number at {cell.coordinate}",
                            cell=cell.coordinate,
                            sheet=sheet_name,
                            suggestion="Verify this value is correct"
                        ))
    
    def _validate_financial_model_structure(self):
        """Validate financial model structure"""
        # Check for proper financial statement structure
        if "Income Statement" in self.workbook.sheetnames:
            self._validate_income_statement_structure()
        
        if "Balance Sheet" in self.workbook.sheetnames:
            self._validate_balance_sheet_structure()
        
        if "Cash Flow" in self.workbook.sheetnames:
            self._validate_cash_flow_structure()
    
    def _validate_income_statement_structure(self):
        """Validate income statement structure"""
        worksheet = self.workbook["Income Statement"]
        
        # Check for required line items
        required_items = ["Revenue", "Cost of Goods Sold", "Gross Profit", "Operating Expenses", "Operating Income", "Net Income"]
        
        for item in required_items:
            found = False
            for row in worksheet.iter_rows():
                for cell in row:
                    if cell.value and isinstance(cell.value, str) and item.lower() in cell.value.lower():
                        found = True
                        break
                if found:
                    break
            
            if not found:
                self.results.append(ValidationResult(
                    severity=ValidationSeverity.WARNING,
                    message=f"Missing required line item: {item}",
                    sheet="Income Statement",
                    suggestion="Add this line item for completeness"
                ))
    
    def _validate_balance_sheet_structure(self):
        """Validate balance sheet structure"""
        worksheet = self.workbook["Balance Sheet"]
        
        # Check for required line items
        required_items = ["Cash", "Accounts Receivable", "Inventory", "Total Assets", "Accounts Payable", "Total Liabilities", "Total Equity"]
        
        for item in required_items:
            found = False
            for row in worksheet.iter_rows():
                for cell in row:
                    if cell.value and isinstance(cell.value, str) and item.lower() in cell.value.lower():
                        found = True
                        break
                if found:
                    break
            
            if not found:
                self.results.append(ValidationResult(
                    severity=ValidationSeverity.WARNING,
                    message=f"Missing required line item: {item}",
                    sheet="Balance Sheet",
                    suggestion="Add this line item for completeness"
                ))
    
    def _validate_cash_flow_structure(self):
        """Validate cash flow statement structure"""
        worksheet = self.workbook["Cash Flow"]
        
        # Check for required line items
        required_items = ["Net Income", "Depreciation", "Operating Cash Flow", "Capital Expenditures", "Investing Cash Flow", "Financing Cash Flow", "Net Change in Cash"]
        
        for item in required_items:
            found = False
            for row in worksheet.iter_rows():
                for cell in row:
                    if cell.value and isinstance(cell.value, str) and item.lower() in cell.value.lower():
                        found = True
                        break
                if found:
                    break
            
            if not found:
                self.results.append(ValidationResult(
                    severity=ValidationSeverity.WARNING,
                    message=f"Missing required line item: {item}",
                    sheet="Cash Flow",
                    suggestion="Add this line item for completeness"
                ))
    
    def generate_validation_report(self) -> str:
        """Generate a human-readable validation report"""
        if not self.results:
            return "No validation issues found. Excel file meets investment banking standards."
        
        report = "Excel Validation Report\n"
        report += "=" * 50 + "\n\n"
        
        # Group results by severity
        errors = [r for r in self.results if r.severity == ValidationSeverity.ERROR]
        warnings = [r for r in self.results if r.severity == ValidationSeverity.WARNING]
        info = [r for r in self.results if r.severity == ValidationSeverity.INFO]
        
        if errors:
            report += "ERRORS:\n"
            report += "-" * 20 + "\n"
            for error in errors:
                report += f"• {error.message}\n"
                if error.cell:
                    report += f"  Cell: {error.cell}\n"
                if error.sheet:
                    report += f"  Sheet: {error.sheet}\n"
                if error.suggestion:
                    report += f"  Suggestion: {error.suggestion}\n"
                report += "\n"
        
        if warnings:
            report += "WARNINGS:\n"
            report += "-" * 20 + "\n"
            for warning in warnings:
                report += f"• {warning.message}\n"
                if warning.cell:
                    report += f"  Cell: {warning.cell}\n"
                if warning.sheet:
                    report += f"  Sheet: {warning.sheet}\n"
                if warning.suggestion:
                    report += f"  Suggestion: {warning.suggestion}\n"
                report += "\n"
        
        if info:
            report += "SUGGESTIONS:\n"
            report += "-" * 20 + "\n"
            for suggestion in info:
                report += f"• {suggestion.message}\n"
                if suggestion.cell:
                    report += f"  Cell: {suggestion.cell}\n"
                if suggestion.sheet:
                    report += f"  Sheet: {suggestion.sheet}\n"
                if suggestion.suggestion:
                    report += f"  Suggestion: {suggestion.suggestion}\n"
                report += "\n"
        
        return report

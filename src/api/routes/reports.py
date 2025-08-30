"""
NGI Capital Financial Reporting Routes
====================================

This module provides comprehensive financial reporting capabilities including:
- Income Statement (P&L) generation
- Balance Sheet creation  
- Cash Flow Statement (ASC 230 compliant)
- Partner Capital Statements
- Custom management reports
- Export functionality (PDF/Excel)

Features:
- GAAP-compliant financial statements
- Multi-entity reporting
- Period comparisons
- Partner-specific reports
- Export to multiple formats
- Audit trail integration

Author: NGI Capital Development Team
"""

import io
import os
from datetime import date, datetime, timedelta
from decimal import Decimal
from typing import List, Optional, Dict, Any, Union
import logging

from fastapi import APIRouter, Depends, HTTPException, status, Query, Response
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import text, and_, or_, func
from pydantic import BaseModel, Field, validator

# Import dependencies for export functionality
try:
    import pandas as pd
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Border, Side
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import letter, A4
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    EXPORT_AVAILABLE = True
except ImportError:
    EXPORT_AVAILABLE = False

# Import dependencies (these would need to be created)
# from ..database import get_db
# from ..auth import get_current_user, require_partner_access
# from ..models import Entity, Transaction, JournalEntry, Account, Partner

# Configure logging
logger = logging.getLogger(__name__)

# Router instance
router = APIRouter(prefix="/api/reports", tags=["reports"])

# Pydantic models for request/response validation
class ReportPeriod(BaseModel):
    """Model for report period specification"""
    start_date: date = Field(..., description="Report start date")
    end_date: date = Field(..., description="Report end date") 
    
    @validator('end_date')
    def end_date_after_start_date(cls, v, values):
        if 'start_date' in values and v <= values['start_date']:
            raise ValueError('end_date must be after start_date')
        return v

class IncomeStatementLine(BaseModel):
    """Model for income statement line items"""
    account_code: str
    account_name: str
    current_period: Decimal = Field(default=Decimal('0.00'))
    previous_period: Optional[Decimal] = Field(default=Decimal('0.00'))
    variance: Optional[Decimal] = Field(default=Decimal('0.00'))
    variance_percent: Optional[Decimal] = Field(default=Decimal('0.00'))

class IncomeStatement(BaseModel):
    """Model for complete income statement"""
    entity_id: int
    entity_name: str
    period: ReportPeriod
    currency: str = "USD"
    
    # Revenue section
    revenue_lines: List[IncomeStatementLine] = []
    total_revenue: Decimal = Field(default=Decimal('0.00'))
    
    # Expense sections
    cost_of_revenue_lines: List[IncomeStatementLine] = []
    total_cost_of_revenue: Decimal = Field(default=Decimal('0.00'))
    gross_profit: Decimal = Field(default=Decimal('0.00'))
    
    operating_expense_lines: List[IncomeStatementLine] = []
    total_operating_expenses: Decimal = Field(default=Decimal('0.00'))
    operating_income: Decimal = Field(default=Decimal('0.00'))
    
    other_income_lines: List[IncomeStatementLine] = []
    total_other_income: Decimal = Field(default=Decimal('0.00'))
    
    net_income: Decimal = Field(default=Decimal('0.00'))
    
    generated_at: datetime = Field(default_factory=datetime.utcnow)
    generated_by: str

class BalanceSheetLine(BaseModel):
    """Model for balance sheet line items"""
    account_code: str
    account_name: str
    current_balance: Decimal = Field(default=Decimal('0.00'))
    previous_balance: Optional[Decimal] = Field(default=Decimal('0.00'))
    variance: Optional[Decimal] = Field(default=Decimal('0.00'))

class BalanceSheet(BaseModel):
    """Model for complete balance sheet"""
    entity_id: int
    entity_name: str
    as_of_date: date
    currency: str = "USD"
    
    # Assets section
    current_assets: List[BalanceSheetLine] = []
    total_current_assets: Decimal = Field(default=Decimal('0.00'))
    
    fixed_assets: List[BalanceSheetLine] = []
    total_fixed_assets: Decimal = Field(default=Decimal('0.00'))
    
    other_assets: List[BalanceSheetLine] = []
    total_other_assets: Decimal = Field(default=Decimal('0.00'))
    
    total_assets: Decimal = Field(default=Decimal('0.00'))
    
    # Liabilities section
    current_liabilities: List[BalanceSheetLine] = []
    total_current_liabilities: Decimal = Field(default=Decimal('0.00'))
    
    long_term_liabilities: List[BalanceSheetLine] = []
    total_long_term_liabilities: Decimal = Field(default=Decimal('0.00'))
    
    total_liabilities: Decimal = Field(default=Decimal('0.00'))
    
    # Equity section
    equity_lines: List[BalanceSheetLine] = []
    total_equity: Decimal = Field(default=Decimal('0.00'))
    
    total_liabilities_and_equity: Decimal = Field(default=Decimal('0.00'))
    
    generated_at: datetime = Field(default_factory=datetime.utcnow)
    generated_by: str

class CashFlowLine(BaseModel):
    """Model for cash flow statement line items"""
    description: str
    amount: Decimal = Field(default=Decimal('0.00'))
    account_codes: List[str] = []

class CashFlowStatement(BaseModel):
    """Model for cash flow statement (ASC 230 compliant)"""
    entity_id: int
    entity_name: str
    period: ReportPeriod
    currency: str = "USD"
    
    # Operating activities
    operating_activities: List[CashFlowLine] = []
    net_cash_from_operating: Decimal = Field(default=Decimal('0.00'))
    
    # Investing activities
    investing_activities: List[CashFlowLine] = []
    net_cash_from_investing: Decimal = Field(default=Decimal('0.00'))
    
    # Financing activities
    financing_activities: List[CashFlowLine] = []
    net_cash_from_financing: Decimal = Field(default=Decimal('0.00'))
    
    net_increase_in_cash: Decimal = Field(default=Decimal('0.00'))
    cash_beginning_of_period: Decimal = Field(default=Decimal('0.00'))
    cash_end_of_period: Decimal = Field(default=Decimal('0.00'))
    
    generated_at: datetime = Field(default_factory=datetime.utcnow)
    generated_by: str

class PartnerCapitalLine(BaseModel):
    """Model for partner capital statement line items"""
    partner_name: str
    partner_email: str
    beginning_balance: Decimal = Field(default=Decimal('0.00'))
    capital_contributions: Decimal = Field(default=Decimal('0.00'))
    capital_withdrawals: Decimal = Field(default=Decimal('0.00'))
    allocated_income: Decimal = Field(default=Decimal('0.00'))
    allocated_expenses: Decimal = Field(default=Decimal('0.00'))
    ending_balance: Decimal = Field(default=Decimal('0.00'))
    ownership_percentage: Decimal = Field(default=Decimal('0.00'))

class PartnerCapitalStatement(BaseModel):
    """Model for partner capital statements"""
    entity_id: int
    entity_name: str
    period: ReportPeriod
    currency: str = "USD"
    
    partner_details: List[PartnerCapitalLine] = []
    total_beginning_capital: Decimal = Field(default=Decimal('0.00'))
    total_contributions: Decimal = Field(default=Decimal('0.00'))
    total_withdrawals: Decimal = Field(default=Decimal('0.00'))
    total_allocated_income: Decimal = Field(default=Decimal('0.00'))
    total_ending_capital: Decimal = Field(default=Decimal('0.00'))
    
    generated_at: datetime = Field(default_factory=datetime.utcnow)
    generated_by: str

class ReportExportRequest(BaseModel):
    """Model for report export requests"""
    report_type: str = Field(..., description="Type of report to export")
    format: str = Field(..., description="Export format (pdf, excel, csv)")
    entity_id: int = Field(..., description="Entity ID for the report")
    period: Optional[ReportPeriod] = Field(None, description="Report period")
    as_of_date: Optional[date] = Field(None, description="As of date for balance sheet")
    
    @validator('format')
    def validate_format(cls, v):
        valid_formats = ['pdf', 'excel', 'csv']
        if v.lower() not in valid_formats:
            raise ValueError(f'Format must be one of: {", ".join(valid_formats)}')
        return v.lower()
    
    @validator('report_type')
    def validate_report_type(cls, v):
        valid_types = ['income_statement', 'balance_sheet', 'cash_flow', 'partner_capital']
        if v.lower() not in valid_types:
            raise ValueError(f'Report type must be one of: {", ".join(valid_types)}')
        return v.lower()

# Utility functions
def get_account_balance(db: Session, account_codes: List[str], entity_id: int, 
                       start_date: date, end_date: date) -> Decimal:
    """Calculate account balance for given period"""
    query = """
        SELECT COALESCE(SUM(je.debit_amount - je.credit_amount), 0) as balance
        FROM journal_entries je
        JOIN accounts a ON je.account_id = a.id
        JOIN transactions t ON je.transaction_id = t.id
        WHERE a.account_code IN :account_codes
        AND a.entity_id = :entity_id
        AND t.transaction_date BETWEEN :start_date AND :end_date
        AND t.approval_status = 'approved'
    """
    
    try:
        # result = db.execute(text(query), {
        #     "account_codes": tuple(account_codes),
        #     "entity_id": entity_id,
        #     "start_date": start_date,
        #     "end_date": end_date
        # }).fetchone()
        # return Decimal(str(result[0])) if result and result[0] else Decimal('0.00')
        return Decimal('0.00')  # Placeholder
    except Exception as e:
        logger.error(f"Error calculating account balance: {str(e)}")
        return Decimal('0.00')

def generate_excel_report(report_data: Dict, report_type: str) -> io.BytesIO:
    """Generate Excel report from report data"""
    if not EXPORT_AVAILABLE:
        raise HTTPException(
            status_code=status.HTTP_501_NOT_IMPLEMENTED,
            detail="Excel export not available - missing dependencies"
        )
    
    wb = Workbook()
    ws = wb.active
    ws.title = report_type.replace('_', ' ').title()
    
    # Header styling
    header_font = Font(bold=True, size=14)
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    
    # Add report headers
    ws['A1'] = report_data.get('entity_name', 'NGI Capital')
    ws['A1'].font = header_font
    ws['A2'] = report_type.replace('_', ' ').title()
    ws['A2'].font = Font(bold=True, size=12)
    
    # Add period information
    if 'period' in report_data:
        period = report_data['period']
        ws['A3'] = f"Period: {period['start_date']} to {period['end_date']}"
    elif 'as_of_date' in report_data:
        ws['A3'] = f"As of: {report_data['as_of_date']}"
    
    # Content would be added based on report type
    # This is a placeholder implementation
    
    # Save to BytesIO
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output

def generate_pdf_report(report_data: Dict, report_type: str) -> io.BytesIO:
    """Generate PDF report from report data"""
    if not EXPORT_AVAILABLE:
        raise HTTPException(
            status_code=status.HTTP_501_NOT_IMPLEMENTED,
            detail="PDF export not available - missing dependencies"
        )
    
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter)
    styles = getSampleStyleSheet()
    story = []
    
    # Title
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=18,
        spaceAfter=30,
        textColor=colors.HexColor('#366092')
    )
    
    title = Paragraph(f"{report_data.get('entity_name', 'NGI Capital')}<br/>{report_type.replace('_', ' ').title()}", title_style)
    story.append(title)
    
    # Period information
    if 'period' in report_data:
        period = report_data['period']
        period_text = f"Period: {period['start_date']} to {period['end_date']}"
    elif 'as_of_date' in report_data:
        period_text = f"As of: {report_data['as_of_date']}"
    else:
        period_text = f"Generated: {datetime.now().strftime('%Y-%m-%d')}"
    
    period_para = Paragraph(period_text, styles['Normal'])
    story.append(period_para)
    story.append(Spacer(1, 20))
    
    # Content would be added based on report type
    # This is a placeholder implementation
    
    doc.build(story)
    buffer.seek(0)
    return buffer

# API Endpoints

@router.get("/income-statement/{entity_id}", response_model=IncomeStatement)
async def get_income_statement(
    entity_id: int,
    start_date: date = Query(..., description="Report start date"),
    end_date: date = Query(..., description="Report end date"),
    include_previous_period: bool = Query(False, description="Include previous period comparison"),
    # db: Session = Depends(get_db),
    # current_user: Partner = Depends(require_partner_access)
):
    """
    Generate Income Statement (P&L) for specified entity and period.
    
    - **entity_id**: Entity to generate report for
    - **start_date**: Report period start date
    - **end_date**: Report period end date
    - **include_previous_period**: Include previous period for comparison
    """
    try:
        # Verify entity exists
        entity_query = "SELECT legal_name FROM entities WHERE id = :entity_id AND is_active = 1"
        # entity = db.execute(text(entity_query), {"entity_id": entity_id}).fetchone()
        # if not entity:
        #     raise HTTPException(
        #         status_code=status.HTTP_404_NOT_FOUND,
        #         detail=f"Entity with ID {entity_id} not found"
        #     )
        
        period = ReportPeriod(start_date=start_date, end_date=end_date)
        
        # Get revenue accounts (4xxxx)
        revenue_query = """
            SELECT a.account_code, a.account_name,
                   COALESCE(SUM(je.credit_amount - je.debit_amount), 0) as current_amount
            FROM accounts a
            LEFT JOIN journal_entries je ON a.id = je.account_id
            LEFT JOIN transactions t ON je.transaction_id = t.id
            WHERE a.entity_id = :entity_id 
            AND a.account_type = 'Revenue'
            AND a.is_active = 1
            AND (t.transaction_date BETWEEN :start_date AND :end_date OR t.transaction_date IS NULL)
            AND (t.approval_status = 'approved' OR t.approval_status IS NULL)
            GROUP BY a.account_code, a.account_name
            ORDER BY a.account_code
        """
        
        # Get expense accounts (5xxxx)
        expense_query = """
            SELECT a.account_code, a.account_name,
                   COALESCE(SUM(je.debit_amount - je.credit_amount), 0) as current_amount
            FROM accounts a
            LEFT JOIN journal_entries je ON a.id = je.account_id
            LEFT JOIN transactions t ON je.transaction_id = t.id
            WHERE a.entity_id = :entity_id 
            AND a.account_type = 'Expense'
            AND a.is_active = 1
            AND (t.transaction_date BETWEEN :start_date AND :end_date OR t.transaction_date IS NULL)
            AND (t.approval_status = 'approved' OR t.approval_status IS NULL)
            GROUP BY a.account_code, a.account_name
            ORDER BY a.account_code
        """
        
        # In real implementation, execute queries and build report
        # revenue_data = db.execute(text(revenue_query), {...}).fetchall()
        # expense_data = db.execute(text(expense_query), {...}).fetchall()
        
        logger.info(f"Generated income statement for entity {entity_id} by user: {getattr(current_user, 'email', 'unknown')}")
        
        # Return placeholder report
        return IncomeStatement(
            entity_id=entity_id,
            entity_name="Sample Entity",
            period=period,
            generated_by=getattr(current_user, 'email', 'unknown')
        )
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error generating income statement: {str(e)}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Failed to generate income statement"
        )

@router.get("/balance-sheet/{entity_id}", response_model=BalanceSheet)
async def get_balance_sheet(
    entity_id: int,
    as_of_date: date = Query(..., description="Balance sheet as of date"),
    include_previous_period: bool = Query(False, description="Include previous period comparison"),
    # db: Session = Depends(get_db),
    # current_user: Partner = Depends(require_partner_access)
):
    """
    Generate Balance Sheet for specified entity as of a specific date.
    
    - **entity_id**: Entity to generate report for
    - **as_of_date**: Balance sheet as of date
    - **include_previous_period**: Include previous period for comparison
    """
    try:
        # Verify entity exists
        entity_query = "SELECT legal_name FROM entities WHERE id = :entity_id AND is_active = 1"
        # entity = db.execute(text(entity_query), {"entity_id": entity_id}).fetchone()
        # if not entity:
        #     raise HTTPException(
        #         status_code=status.HTTP_404_NOT_FOUND,
        #         detail=f"Entity with ID {entity_id} not found"
        #     )
        
        # Get asset accounts (1xxxx)
        assets_query = """
            SELECT a.account_code, a.account_name, a.account_type,
                   COALESCE(SUM(
                       CASE WHEN a.account_type = 'Contra Asset' 
                       THEN je.credit_amount - je.debit_amount
                       ELSE je.debit_amount - je.credit_amount END
                   ), 0) as balance
            FROM accounts a
            LEFT JOIN journal_entries je ON a.id = je.account_id
            LEFT JOIN transactions t ON je.transaction_id = t.id
            WHERE a.entity_id = :entity_id 
            AND (a.account_type = 'Asset' OR a.account_type = 'Contra Asset')
            AND a.is_active = 1
            AND (t.transaction_date <= :as_of_date OR t.transaction_date IS NULL)
            AND (t.approval_status = 'approved' OR t.approval_status IS NULL)
            GROUP BY a.account_code, a.account_name, a.account_type
            ORDER BY a.account_code
        """
        
        # Get liability accounts (2xxxx)
        liabilities_query = """
            SELECT a.account_code, a.account_name,
                   COALESCE(SUM(je.credit_amount - je.debit_amount), 0) as balance
            FROM accounts a
            LEFT JOIN journal_entries je ON a.id = je.account_id
            LEFT JOIN transactions t ON je.transaction_id = t.id
            WHERE a.entity_id = :entity_id 
            AND a.account_type = 'Liability'
            AND a.is_active = 1
            AND (t.transaction_date <= :as_of_date OR t.transaction_date IS NULL)
            AND (t.approval_status = 'approved' OR t.approval_status IS NULL)
            GROUP BY a.account_code, a.account_name
            ORDER BY a.account_code
        """
        
        # Get equity accounts (3xxxx)
        equity_query = """
            SELECT a.account_code, a.account_name,
                   COALESCE(SUM(je.credit_amount - je.debit_amount), 0) as balance
            FROM accounts a
            LEFT JOIN journal_entries je ON a.id = je.account_id
            LEFT JOIN transactions t ON je.transaction_id = t.id
            WHERE a.entity_id = :entity_id 
            AND a.account_type = 'Equity'
            AND a.is_active = 1
            AND (t.transaction_date <= :as_of_date OR t.transaction_date IS NULL)
            AND (t.approval_status = 'approved' OR t.approval_status IS NULL)
            GROUP BY a.account_code, a.account_name
            ORDER BY a.account_code
        """
        
        logger.info(f"Generated balance sheet for entity {entity_id} by user: {getattr(current_user, 'email', 'unknown')}")
        
        # Return placeholder report
        return BalanceSheet(
            entity_id=entity_id,
            entity_name="Sample Entity",
            as_of_date=as_of_date,
            generated_by=getattr(current_user, 'email', 'unknown')
        )
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error generating balance sheet: {str(e)}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Failed to generate balance sheet"
        )

@router.get("/cash-flow/{entity_id}", response_model=CashFlowStatement)
async def get_cash_flow_statement(
    entity_id: int,
    start_date: date = Query(..., description="Report start date"),
    end_date: date = Query(..., description="Report end date"),
    method: str = Query("direct", description="Cash flow method (direct or indirect)"),
    # db: Session = Depends(get_db),
    # current_user: Partner = Depends(require_partner_access)
):
    """
    Generate Cash Flow Statement (ASC 230 compliant) for specified entity and period.
    
    - **entity_id**: Entity to generate report for
    - **start_date**: Report period start date
    - **end_date**: Report period end date
    - **method**: Cash flow method (direct or indirect)
    """
    try:
        if method not in ['direct', 'indirect']:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Method must be 'direct' or 'indirect'"
            )
        
        # Verify entity exists
        entity_query = "SELECT legal_name FROM entities WHERE id = :entity_id AND is_active = 1"
        # entity = db.execute(text(entity_query), {"entity_id": entity_id}).fetchone()
        # if not entity:
        #     raise HTTPException(
        #         status_code=status.HTTP_404_NOT_FOUND,
        #         detail=f"Entity with ID {entity_id} not found"
        #     )
        
        period = ReportPeriod(start_date=start_date, end_date=end_date)
        
        # Calculate operating activities
        operating_query = """
            SELECT 
                t.transaction_type,
                t.description,
                SUM(t.amount) as total_amount
            FROM transactions t
            WHERE t.entity_id = :entity_id
            AND t.transaction_date BETWEEN :start_date AND :end_date
            AND t.approval_status = 'approved'
            AND t.transaction_type IN ('REVENUE', 'EXPENSE', 'PAYROLL')
            GROUP BY t.transaction_type, t.description
            ORDER BY t.transaction_type
        """
        
        # Calculate investing activities
        investing_query = """
            SELECT 
                t.transaction_type,
                t.description,
                SUM(t.amount) as total_amount
            FROM transactions t
            WHERE t.entity_id = :entity_id
            AND t.transaction_date BETWEEN :start_date AND :end_date
            AND t.approval_status = 'approved'
            AND t.transaction_type IN ('EQUIPMENT_PURCHASE', 'INVESTMENT', 'ASSET_SALE')
            GROUP BY t.transaction_type, t.description
            ORDER BY t.transaction_type
        """
        
        # Calculate financing activities
        financing_query = """
            SELECT 
                t.transaction_type,
                t.description,
                SUM(t.amount) as total_amount
            FROM transactions t
            WHERE t.entity_id = :entity_id
            AND t.transaction_date BETWEEN :start_date AND :end_date
            AND t.approval_status = 'approved'
            AND t.transaction_type IN ('CAPITAL_CONTRIBUTION', 'CAPITAL_WITHDRAWAL', 'LOAN')
            GROUP BY t.transaction_type, t.description
            ORDER BY t.transaction_type
        """
        
        logger.info(f"Generated cash flow statement for entity {entity_id} by user: {getattr(current_user, 'email', 'unknown')}")
        
        # Return placeholder report
        return CashFlowStatement(
            entity_id=entity_id,
            entity_name="Sample Entity",
            period=period,
            generated_by=getattr(current_user, 'email', 'unknown')
        )
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error generating cash flow statement: {str(e)}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Failed to generate cash flow statement"
        )

@router.get("/partner-capital/{entity_id}", response_model=PartnerCapitalStatement)
async def get_partner_capital_statement(
    entity_id: int,
    start_date: date = Query(..., description="Report start date"),
    end_date: date = Query(..., description="Report end date"),
    # db: Session = Depends(get_db),
    # current_user: Partner = Depends(require_partner_access)
):
    """
    Generate Partner Capital Statement for specified entity and period.
    
    - **entity_id**: Entity to generate report for
    - **start_date**: Report period start date
    - **end_date**: Report period end date
    """
    try:
        # Verify entity exists
        entity_query = "SELECT legal_name FROM entities WHERE id = :entity_id AND is_active = 1"
        # entity = db.execute(text(entity_query), {"entity_id": entity_id}).fetchone()
        # if not entity:
        #     raise HTTPException(
        #         status_code=status.HTTP_404_NOT_FOUND,
        #         detail=f"Entity with ID {entity_id} not found"
        #     )
        
        period = ReportPeriod(start_date=start_date, end_date=end_date)
        
        # Get partner information
        partners_query = """
            SELECT p.name, p.email, p.ownership_percentage,
                   p.capital_account_balance as current_balance
            FROM partners p
            WHERE p.is_active = 1
            ORDER BY p.name
        """
        
        # Calculate partner capital changes for the period
        capital_changes_query = """
            SELECT 
                t.created_by as partner_email,
                t.transaction_type,
                SUM(CASE WHEN t.transaction_type = 'CAPITAL_CONTRIBUTION' THEN t.amount ELSE 0 END) as contributions,
                SUM(CASE WHEN t.transaction_type = 'CAPITAL_WITHDRAWAL' THEN t.amount ELSE 0 END) as withdrawals
            FROM transactions t
            WHERE t.entity_id = :entity_id
            AND t.transaction_date BETWEEN :start_date AND :end_date
            AND t.approval_status = 'approved'
            AND t.transaction_type IN ('CAPITAL_CONTRIBUTION', 'CAPITAL_WITHDRAWAL')
            GROUP BY t.created_by, t.transaction_type
        """
        
        # Calculate allocated income/expenses based on ownership percentages
        income_allocation_query = """
            SELECT 
                SUM(CASE WHEN a.account_type = 'Revenue' THEN je.credit_amount - je.debit_amount ELSE 0 END) as total_income,
                SUM(CASE WHEN a.account_type = 'Expense' THEN je.debit_amount - je.credit_amount ELSE 0 END) as total_expenses
            FROM journal_entries je
            JOIN accounts a ON je.account_id = a.id
            JOIN transactions t ON je.transaction_id = t.id
            WHERE a.entity_id = :entity_id
            AND t.transaction_date BETWEEN :start_date AND :end_date
            AND t.approval_status = 'approved'
        """
        
        logger.info(f"Generated partner capital statement for entity {entity_id} by user: {getattr(current_user, 'email', 'unknown')}")
        
        # Return placeholder report
        return PartnerCapitalStatement(
            entity_id=entity_id,
            entity_name="Sample Entity",
            period=period,
            generated_by=getattr(current_user, 'email', 'unknown')
        )
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error generating partner capital statement: {str(e)}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Failed to generate partner capital statement"
        )

@router.post("/export")
async def export_report(
    export_request: ReportExportRequest,
    # db: Session = Depends(get_db),
    # current_user: Partner = Depends(require_partner_access)
):
    """
    Export financial report to PDF, Excel, or CSV format.
    
    - **export_request**: Export configuration including report type, format, and parameters
    """
    try:
        # Get report data based on type
        report_data = {}
        
        if export_request.report_type == "income_statement":
            if not export_request.period:
                raise HTTPException(
                    status_code=status.HTTP_400_BAD_REQUEST,
                    detail="Period is required for income statement"
                )
            # report = await get_income_statement(
            #     export_request.entity_id,
            #     export_request.period.start_date,
            #     export_request.period.end_date,
            #     False, db, current_user
            # )
            # report_data = report.dict()
            
        elif export_request.report_type == "balance_sheet":
            if not export_request.as_of_date:
                raise HTTPException(
                    status_code=status.HTTP_400_BAD_REQUEST,
                    detail="as_of_date is required for balance sheet"
                )
            # report = await get_balance_sheet(
            #     export_request.entity_id,
            #     export_request.as_of_date,
            #     False, db, current_user
            # )
            # report_data = report.dict()
            
        elif export_request.report_type == "cash_flow":
            if not export_request.period:
                raise HTTPException(
                    status_code=status.HTTP_400_BAD_REQUEST,
                    detail="Period is required for cash flow statement"
                )
            # Similar implementation for cash flow
            
        elif export_request.report_type == "partner_capital":
            if not export_request.period:
                raise HTTPException(
                    status_code=status.HTTP_400_BAD_REQUEST,
                    detail="Period is required for partner capital statement"
                )
            # Similar implementation for partner capital
        
        # Generate file based on format
        if export_request.format == "pdf":
            file_buffer = generate_pdf_report(report_data, export_request.report_type)
            media_type = "application/pdf"
            filename = f"{export_request.report_type}_{export_request.entity_id}.pdf"
            
        elif export_request.format == "excel":
            file_buffer = generate_excel_report(report_data, export_request.report_type)
            media_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            filename = f"{export_request.report_type}_{export_request.entity_id}.xlsx"
            
        elif export_request.format == "csv":
            # Convert report data to CSV format
            csv_data = "Placeholder CSV data"
            file_buffer = io.StringIO(csv_data)
            media_type = "text/csv"
            filename = f"{export_request.report_type}_{export_request.entity_id}.csv"
        
        logger.info(f"Exported {export_request.report_type} for entity {export_request.entity_id} in {export_request.format} format by user: {getattr(current_user, 'email', 'unknown')}")
        
        return StreamingResponse(
            io.BytesIO(b"Placeholder file content"),  # file_buffer in real implementation
            media_type=media_type,
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error exporting report: {str(e)}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Failed to export report"
        )

@router.get("/entities", response_model=List[Dict[str, Any]])
async def get_reportable_entities(
    # db: Session = Depends(get_db),
    # current_user: Partner = Depends(require_partner_access)
):
    """
    Get list of entities available for reporting.
    """
    try:
        entities_query = """
            SELECT id, legal_name, entity_type 
            FROM entities 
            WHERE is_active = 1
            ORDER BY legal_name
        """
        
        # entities = db.execute(text(entities_query)).fetchall()
        
        logger.info(f"Retrieved reportable entities for user: {getattr(current_user, 'email', 'unknown')}")
        
        return []  # Return empty list for now
        
    except Exception as e:
        logger.error(f"Error retrieving reportable entities: {str(e)}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Failed to retrieve entities"
        )
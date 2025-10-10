"""
Big 4 Audit Package Generator
Generates comprehensive fixed asset audit packages per Big 4 requirements
(PwC, Deloitte, EY, KPMG standards)
"""

from decimal import Decimal
from datetime import date
from typing import Dict, List
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, and_, func, extract
from src.api.models_fixed_assets import FixedAsset, DepreciationEntry, AssetDisposal, AuditPackage
from src.api.models_accounting import AccountingEntity
from src.api.utils.datetime_utils import get_pst_now
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os
import logging

logger = logging.getLogger(__name__)


class AuditPackageGenerator:
    """Generates Big 4 audit packages for fixed assets"""
    
    # PwC/Deloitte/EY/KPMG standard colors
    HEADER_FILL = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
    SUBTOTAL_FILL = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    SUBTOTAL_FONT = Font(bold=True, size=10)
    TOTAL_FILL = PatternFill(start_color="305496", end_color="305496", fill_type="solid")
    TOTAL_FONT = Font(bold=True, color="FFFFFF", size=11)
    
    THIN_BORDER = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    @staticmethod
    async def generate_fixed_asset_audit_package(
        entity_id: int,
        year: int,
        user_email: str,
        db: AsyncSession
    ) -> str:
        """
        Generate complete Big 4 audit package for fixed assets
        
        Returns:
            File path to generated Excel workbook
        """
        # Get entity
        entity = await db.get(AccountingEntity, entity_id)
        if not entity:
            raise ValueError(f"Entity {entity_id} not found")
        
        period_start = date(year, 1, 1)
        period_end = date(year, 12, 31)
        
        # Create workbook
        wb = openpyxl.Workbook()
        wb.remove(wb.active)  # Remove default sheet
        
        # Generate all schedules
        await AuditPackageGenerator._create_summary_sheet(wb, entity, year, period_end, db)
        await AuditPackageGenerator._create_asset_register(wb, entity_id, period_end, db)
        await AuditPackageGenerator._create_depreciation_schedule(wb, entity_id, year, db)
        await AuditPackageGenerator._create_roll_forward(wb, entity_id, year, db)
        await AuditPackageGenerator._create_additions_schedule(wb, entity_id, year, db)
        await AuditPackageGenerator._create_disposals_schedule(wb, entity_id, year, db)
        
        # Save workbook
        os.makedirs("uploads/audit_packages", exist_ok=True)
        filename = f"Fixed_Asset_Audit_Package_{entity.entity_name.replace(' ', '_')}_{year}.xlsx"
        filepath = f"uploads/audit_packages/{filename}"
        wb.save(filepath)
        
        # Get file size
        file_size = os.path.getsize(filepath)
        
        # Get statistics
        result = await db.execute(
            select(FixedAsset).where(FixedAsset.entity_id == entity_id)
        )
        assets = result.scalars().all()
        
        total_cost = sum(a.acquisition_cost for a in assets)
        total_accum_dep = sum(a.accumulated_depreciation or Decimal("0") for a in assets)
        total_nbv = total_cost - total_accum_dep
        
        # Create audit package record
        audit_package = AuditPackage(
            entity_id=entity_id,
            package_type="Fixed Assets",
            period_year=year,
            period_start=period_start,
            period_end=period_end,
            file_path=filepath,
            file_name=filename,
            file_size_bytes=file_size,
            includes_asset_register=True,
            includes_depreciation_schedule=True,
            includes_roll_forward=True,
            includes_additions_schedule=True,
            includes_disposals_schedule=True,
            total_assets_count=len(assets),
            total_original_cost=total_cost,
            total_accumulated_depreciation=total_accum_dep,
            total_net_book_value=total_nbv,
            generated_at=get_pst_now(),
            generated_by_email=user_email
        )
        db.add(audit_package)
        await db.commit()
        
        logger.info(f"Generated audit package: {filename}")
        
        return filepath
    
    @staticmethod
    async def _create_summary_sheet(
        wb: openpyxl.Workbook,
        entity: AccountingEntity,
        year: int,
        as_of_date: date,
        db: AsyncSession
    ):
        """Create summary/cover sheet"""
        ws = wb.create_sheet("Summary", 0)
        
        # Title
        ws['A1'] = f"{entity.entity_name}"
        ws['A1'].font = Font(bold=True, size=16)
        ws['A2'] = f"Fixed Asset Audit Package"
        ws['A2'].font = Font(bold=True, size=14)
        ws['A3'] = f"Year Ended December 31, {year}"
        ws['A3'].font = Font(size=12)
        ws['A4'] = f"Generated: {get_pst_now().strftime('%B %d, %Y')}"
        
        # Entity information
        ws['A6'] = "Entity Information"
        ws['A6'].font = Font(bold=True, size=12)
        ws['A7'] = f"Entity Type: {entity.entity_type}"
        ws['A8'] = f"EIN: {entity.ein or 'N/A'}"
        ws['A9'] = f"Formation Date: {entity.formation_date or 'N/A'}"
        
        # Get asset statistics
        result = await db.execute(
            select(FixedAsset).where(FixedAsset.entity_id == entity.id)
        )
        assets = result.scalars().all()
        
        # Summary statistics
        ws['A11'] = "Fixed Assets Summary"
        ws['A11'].font = Font(bold=True, size=12)
        ws['A12'] = "Total Assets:"
        ws['B12'] = len(assets)
        ws['A13'] = "Original Cost:"
        ws['B13'] = float(sum(a.acquisition_cost for a in assets))
        ws['B13'].number_format = '$#,##0.00'
        ws['A14'] = "Accumulated Depreciation:"
        ws['B14'] = float(sum(a.accumulated_depreciation or Decimal("0") for a in assets))
        ws['B14'].number_format = '$#,##0.00'
        ws['A15'] = "Net Book Value:"
        ws['B15'] = float(sum(a.net_book_value or a.acquisition_cost for a in assets))
        ws['B15'].number_format = '$#,##0.00'
        
        # Category breakdown
        ws['A17'] = "Assets by Category"
        ws['A17'].font = Font(bold=True, size=12)
        
        categories = {}
        for asset in assets:
            cat = asset.asset_category
            if cat not in categories:
                categories[cat] = {"count": 0, "cost": Decimal("0"), "nbv": Decimal("0")}
            categories[cat]["count"] += 1
            categories[cat]["cost"] += asset.acquisition_cost
            categories[cat]["nbv"] += asset.net_book_value or asset.acquisition_cost
        
        row = 18
        for cat, data in sorted(categories.items()):
            ws[f'A{row}'] = cat
            ws[f'B{row}'] = data["count"]
            ws[f'C{row}'] = float(data["cost"])
            ws[f'C{row}'].number_format = '$#,##0.00'
            ws[f'D{row}'] = float(data["nbv"])
            ws[f'D{row}'].number_format = '$#,##0.00'
            row += 1
        
        # Workbook contents
        ws[f'A{row + 2}'] = "Workbook Contents"
        ws[f'A{row + 2}'].font = Font(bold=True, size=12)
        ws[f'A{row + 3}'] = "PBC-1: Fixed Asset Register"
        ws[f'A{row + 4}'] = "PBC-2: Depreciation Schedule"
        ws[f'A{row + 5}'] = "PBC-3: Roll Forward Report"
        ws[f'A{row + 6}'] = "PBC-4: Additions Schedule"
        ws[f'A{row + 7}'] = "PBC-5: Disposals Schedule"
        
        # Format columns
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15
    
    @staticmethod
    async def _create_asset_register(
        wb: openpyxl.Workbook,
        entity_id: int,
        as_of_date: date,
        db: AsyncSession
    ):
        """Create PBC-1: Fixed Asset Register"""
        ws = wb.create_sheet("PBC-1 Asset Register")
        
        # Headers
        headers = [
            "Asset Number", "Asset Name", "Category", "Description",
            "Acquisition Date", "Cost", "Salvage Value", "Useful Life (Years)",
            "Depreciation Method", "Accumulated Depreciation", "Net Book Value",
            "Status", "Location"
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.value = header
            cell.font = AuditPackageGenerator.HEADER_FONT
            cell.fill = AuditPackageGenerator.HEADER_FILL
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = AuditPackageGenerator.THIN_BORDER
        
        # Get all assets
        result = await db.execute(
            select(FixedAsset)
            .where(FixedAsset.entity_id == entity_id)
            .order_by(FixedAsset.asset_number)
        )
        assets = result.scalars().all()
        
        # Data rows
        row = 2
        total_cost = Decimal("0")
        total_accum_dep = Decimal("0")
        total_nbv = Decimal("0")
        
        for asset in assets:
            ws.cell(row=row, column=1).value = asset.asset_number
            ws.cell(row=row, column=2).value = asset.asset_name
            ws.cell(row=row, column=3).value = asset.asset_category
            ws.cell(row=row, column=4).value = asset.asset_description or ""
            ws.cell(row=row, column=5).value = asset.acquisition_date
            ws.cell(row=row, column=5).number_format = 'MM/DD/YYYY'
            ws.cell(row=row, column=6).value = float(asset.acquisition_cost)
            ws.cell(row=row, column=6).number_format = '$#,##0.00'
            ws.cell(row=row, column=7).value = float(asset.salvage_value or 0)
            ws.cell(row=row, column=7).number_format = '$#,##0.00'
            ws.cell(row=row, column=8).value = asset.useful_life_years
            ws.cell(row=row, column=9).value = asset.depreciation_method
            ws.cell(row=row, column=10).value = float(asset.accumulated_depreciation or 0)
            ws.cell(row=row, column=10).number_format = '$#,##0.00'
            ws.cell(row=row, column=11).value = float(asset.net_book_value or asset.acquisition_cost)
            ws.cell(row=row, column=11).number_format = '$#,##0.00'
            ws.cell(row=row, column=12).value = asset.status
            ws.cell(row=row, column=13).value = asset.location or ""
            
            total_cost += asset.acquisition_cost
            total_accum_dep += asset.accumulated_depreciation or Decimal("0")
            total_nbv += asset.net_book_value or asset.acquisition_cost
            
            row += 1
        
        # Totals row
        ws.cell(row=row, column=1).value = "TOTAL"
        ws.cell(row=row, column=1).font = AuditPackageGenerator.TOTAL_FONT
        ws.cell(row=row, column=6).value = float(total_cost)
        ws.cell(row=row, column=6).number_format = '$#,##0.00'
        ws.cell(row=row, column=6).font = AuditPackageGenerator.TOTAL_FONT
        ws.cell(row=row, column=10).value = float(total_accum_dep)
        ws.cell(row=row, column=10).number_format = '$#,##0.00'
        ws.cell(row=row, column=10).font = AuditPackageGenerator.TOTAL_FONT
        ws.cell(row=row, column=11).value = float(total_nbv)
        ws.cell(row=row, column=11).number_format = '$#,##0.00'
        ws.cell(row=row, column=11).font = AuditPackageGenerator.TOTAL_FONT
        
        # Format columns
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 35
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 15
        ws.column_dimensions['G'].width = 15
        ws.column_dimensions['H'].width = 15
        ws.column_dimensions['I'].width = 20
        ws.column_dimensions['J'].width = 20
        ws.column_dimensions['K'].width = 15
        ws.column_dimensions['L'].width = 15
        ws.column_dimensions['M'].width = 20
        
        # Freeze panes
        ws.freeze_panes = 'A2'
    
    @staticmethod
    async def _create_depreciation_schedule(
        wb: openpyxl.Workbook,
        entity_id: int,
        year: int,
        db: AsyncSession
    ):
        """Create PBC-2: Depreciation Schedule"""
        ws = wb.create_sheet("PBC-2 Depreciation")
        
        # Headers
        headers = [
            "Asset Number", "Asset Name", "Acquisition Date", "Cost",
            "Jan", "Feb", "Mar", "Apr", "May", "Jun",
            "Jul", "Aug", "Sep", "Oct", "Nov", "Dec",
            "Total Depreciation", "Accumulated Dep", "Net Book Value"
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.value = header
            cell.font = AuditPackageGenerator.HEADER_FONT
            cell.fill = AuditPackageGenerator.HEADER_FILL
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = AuditPackageGenerator.THIN_BORDER
        
        # Get all assets
        result = await db.execute(
            select(FixedAsset)
            .where(FixedAsset.entity_id == entity_id)
            .order_by(FixedAsset.asset_number)
        )
        assets = result.scalars().all()
        
        row = 2
        for asset in assets:
            ws.cell(row=row, column=1).value = asset.asset_number
            ws.cell(row=row, column=2).value = asset.asset_name
            ws.cell(row=row, column=3).value = asset.acquisition_date
            ws.cell(row=row, column=3).number_format = 'MM/DD/YYYY'
            ws.cell(row=row, column=4).value = float(asset.acquisition_cost)
            ws.cell(row=row, column=4).number_format = '$#,##0.00'
            
            # Get monthly depreciation entries for this asset for the year
            dep_result = await db.execute(
                select(DepreciationEntry)
                .where(
                    and_(
                        DepreciationEntry.asset_id == asset.id,
                        DepreciationEntry.period_year == year
                    )
                )
                .order_by(DepreciationEntry.period_month)
            )
            dep_entries = dep_result.scalars().all()
            
            # Map depreciation to months
            monthly_dep = {entry.period_month: entry.depreciation_amount for entry in dep_entries}
            
            # Fill in monthly columns
            total_year_dep = Decimal("0")
            for month in range(1, 13):
                col = 4 + month
                dep_amount = monthly_dep.get(month, Decimal("0"))
                ws.cell(row=row, column=col).value = float(dep_amount)
                ws.cell(row=row, column=col).number_format = '$#,##0'
                total_year_dep += dep_amount
            
            # Total depreciation
            ws.cell(row=row, column=17).value = float(total_year_dep)
            ws.cell(row=row, column=17).number_format = '$#,##0.00'
            
            # Accumulated depreciation
            ws.cell(row=row, column=18).value = float(asset.accumulated_depreciation or 0)
            ws.cell(row=row, column=18).number_format = '$#,##0.00'
            
            # Net book value
            ws.cell(row=row, column=19).value = float(asset.net_book_value or asset.acquisition_cost)
            ws.cell(row=row, column=19).number_format = '$#,##0.00'
            
            row += 1
        
        # Format columns
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 12
        for col in range(5, 17):
            ws.column_dimensions[get_column_letter(col)].width = 10
        ws.column_dimensions['Q'].width = 15
        ws.column_dimensions['R'].width = 18
        ws.column_dimensions['S'].width = 15
        
        ws.freeze_panes = 'A2'
    
    @staticmethod
    async def _create_roll_forward(
        wb: openpyxl.Workbook,
        entity_id: int,
        year: int,
        db: AsyncSession
    ):
        """Create PBC-3: Roll Forward Report"""
        ws = wb.create_sheet("PBC-3 Roll Forward")
        
        # Title
        ws['A1'] = f"Fixed Assets Roll Forward - Year {year}"
        ws['A1'].font = Font(bold=True, size=14)
        ws.merge_cells('A1:F1')
        
        # Headers
        headers = ["", "Beginning Balance", "Additions", "Disposals", "Depreciation", "Ending Balance"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col)
            cell.value = header
            cell.font = AuditPackageGenerator.HEADER_FONT
            cell.fill = AuditPackageGenerator.HEADER_FILL
            cell.alignment = Alignment(horizontal='center')
        
        # Get assets as of beginning of year
        beg_date = date(year - 1, 12, 31)
        end_date = date(year, 12, 31)
        
        # Calculate beginning balances (would need historical data)
        # For now, use current less this year's activity
        
        result = await db.execute(
            select(FixedAsset).where(FixedAsset.entity_id == entity_id)
        )
        assets = result.scalars().all()
        
        # Get additions in year
        additions_result = await db.execute(
            select(FixedAsset).where(
                and_(
                    FixedAsset.entity_id == entity_id,
                    extract('year', FixedAsset.acquisition_date) == year
                )
            )
        )
        additions = additions_result.scalars().all()
        additions_cost = sum(a.acquisition_cost for a in additions)
        
        # Get disposals in year
        disposals_result = await db.execute(
            select(AssetDisposal).where(
                and_(
                    AssetDisposal.entity_id == entity_id,
                    extract('year', AssetDisposal.disposal_date) == year
                )
            )
        )
        disposals = disposals_result.scalars().all()
        disposals_cost = sum(d.original_cost for d in disposals)
        disposals_accum_dep = sum(d.accumulated_depreciation for d in disposals)
        
        # Get depreciation for year
        dep_result = await db.execute(
            select(func.sum(DepreciationEntry.depreciation_amount))
            .where(
                and_(
                    DepreciationEntry.entity_id == entity_id,
                    DepreciationEntry.period_year == year
                )
            )
        )
        year_depreciation = dep_result.scalar() or Decimal("0")
        
        # Calculate balances
        ending_cost = sum(a.acquisition_cost for a in assets)
        ending_accum_dep = sum(a.accumulated_depreciation or Decimal("0") for a in assets)
        ending_nbv = ending_cost - ending_accum_dep
        
        beginning_cost = ending_cost - additions_cost + disposals_cost
        beginning_accum_dep = ending_accum_dep - year_depreciation + disposals_accum_dep
        beginning_nbv = beginning_cost - beginning_accum_dep
        
        # Data rows
        row = 4
        
        # Cost
        ws.cell(row=row, column=1).value = "Cost"
        ws.cell(row=row, column=1).font = Font(bold=True)
        ws.cell(row=row, column=2).value = float(beginning_cost)
        ws.cell(row=row, column=2).number_format = '$#,##0.00'
        ws.cell(row=row, column=3).value = float(additions_cost)
        ws.cell(row=row, column=3).number_format = '$#,##0.00'
        ws.cell(row=row, column=4).value = float(disposals_cost)
        ws.cell(row=row, column=4).number_format = '$(#,##0.00)'
        ws.cell(row=row, column=5).value = 0
        ws.cell(row=row, column=6).value = float(ending_cost)
        ws.cell(row=row, column=6).number_format = '$#,##0.00'
        row += 1
        
        # Accumulated Depreciation
        ws.cell(row=row, column=1).value = "Accumulated Depreciation"
        ws.cell(row=row, column=1).font = Font(bold=True)
        ws.cell(row=row, column=2).value = float(beginning_accum_dep)
        ws.cell(row=row, column=2).number_format = '$(#,##0.00)'
        ws.cell(row=row, column=3).value = 0
        ws.cell(row=row, column=4).value = float(disposals_accum_dep)
        ws.cell(row=row, column=4).number_format = '$#,##0.00'
        ws.cell(row=row, column=5).value = float(year_depreciation)
        ws.cell(row=row, column=5).number_format = '$(#,##0.00)'
        ws.cell(row=row, column=6).value = float(ending_accum_dep)
        ws.cell(row=row, column=6).number_format = '$(#,##0.00)'
        row += 1
        
        # Net Book Value
        ws.cell(row=row, column=1).value = "Net Book Value"
        ws.cell(row=row, column=1).font = AuditPackageGenerator.TOTAL_FONT
        ws.cell(row=row, column=2).value = float(beginning_nbv)
        ws.cell(row=row, column=2).number_format = '$#,##0.00'
        ws.cell(row=row, column=2).font = AuditPackageGenerator.TOTAL_FONT
        ws.cell(row=row, column=3).value = float(additions_cost)
        ws.cell(row=row, column=3).number_format = '$#,##0.00'
        ws.cell(row=row, column=3).font = AuditPackageGenerator.TOTAL_FONT
        ws.cell(row=row, column=4).value = float(disposals_cost - disposals_accum_dep)
        ws.cell(row=row, column=4).number_format = '$(#,##0.00)'
        ws.cell(row=row, column=4).font = AuditPackageGenerator.TOTAL_FONT
        ws.cell(row=row, column=5).value = float(year_depreciation)
        ws.cell(row=row, column=5).number_format = '$(#,##0.00)'
        ws.cell(row=row, column=5).font = AuditPackageGenerator.TOTAL_FONT
        ws.cell(row=row, column=6).value = float(ending_nbv)
        ws.cell(row=row, column=6).number_format = '$#,##0.00'
        ws.cell(row=row, column=6).font = AuditPackageGenerator.TOTAL_FONT
        
        # Format columns
        ws.column_dimensions['A'].width = 25
        for col in ['B', 'C', 'D', 'E', 'F']:
            ws.column_dimensions[col].width = 18
    
    @staticmethod
    async def _create_additions_schedule(
        wb: openpyxl.Workbook,
        entity_id: int,
        year: int,
        db: AsyncSession
    ):
        """Create PBC-4: Additions Schedule"""
        ws = wb.create_sheet("PBC-4 Additions")
        
        # Headers
        headers = [
            "Asset Number", "Asset Name", "Category", "Acquisition Date",
            "Cost", "Vendor", "Auto-Detected", "Confidence %", "JE Number"
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.value = header
            cell.font = AuditPackageGenerator.HEADER_FONT
            cell.fill = AuditPackageGenerator.HEADER_FILL
            cell.alignment = Alignment(horizontal='center')
        
        # Get additions for year
        result = await db.execute(
            select(FixedAsset)
            .where(
                and_(
                    FixedAsset.entity_id == entity_id,
                    extract('year', FixedAsset.acquisition_date) == year
                )
            )
            .order_by(FixedAsset.acquisition_date)
        )
        additions = result.scalars().all()
        
        row = 2
        total_cost = Decimal("0")
        
        for asset in additions:
            ws.cell(row=row, column=1).value = asset.asset_number
            ws.cell(row=row, column=2).value = asset.asset_name
            ws.cell(row=row, column=3).value = asset.asset_category
            ws.cell(row=row, column=4).value = asset.acquisition_date
            ws.cell(row=row, column=4).number_format = 'MM/DD/YYYY'
            ws.cell(row=row, column=5).value = float(asset.acquisition_cost)
            ws.cell(row=row, column=5).number_format = '$#,##0.00'
            ws.cell(row=row, column=6).value = asset.detection_metadata.get("vendor_name", "") if asset.detection_metadata else ""
            ws.cell(row=row, column=7).value = "Yes" if asset.auto_detected else "No"
            ws.cell(row=row, column=8).value = float(asset.detection_confidence or 0)
            ws.cell(row=row, column=8).number_format = '0.0"%"'
            ws.cell(row=row, column=9).value = ""  # Would link to JE if available
            
            total_cost += asset.acquisition_cost
            row += 1
        
        # Total
        ws.cell(row=row, column=1).value = "TOTAL"
        ws.cell(row=row, column=1).font = AuditPackageGenerator.TOTAL_FONT
        ws.cell(row=row, column=5).value = float(total_cost)
        ws.cell(row=row, column=5).number_format = '$#,##0.00'
        ws.cell(row=row, column=5).font = AuditPackageGenerator.TOTAL_FONT
        
        # Format columns
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 25
        ws.column_dimensions['G'].width = 15
        ws.column_dimensions['H'].width = 12
        ws.column_dimensions['I'].width = 15
        
        ws.freeze_panes = 'A2'
    
    @staticmethod
    async def _create_disposals_schedule(
        wb: openpyxl.Workbook,
        entity_id: int,
        year: int,
        db: AsyncSession
    ):
        """Create PBC-5: Disposals Schedule"""
        ws = wb.create_sheet("PBC-5 Disposals")
        
        # Headers
        headers = [
            "Asset Number", "Asset Name", "Disposal Date", "Disposal Type",
            "Original Cost", "Accumulated Dep", "Net Book Value",
            "Disposal Amount", "Gain/(Loss)", "Notes"
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.value = header
            cell.font = AuditPackageGenerator.HEADER_FONT
            cell.fill = AuditPackageGenerator.HEADER_FILL
            cell.alignment = Alignment(horizontal='center')
        
        # Get disposals for year
        result = await db.execute(
            select(AssetDisposal, FixedAsset)
            .join(FixedAsset, AssetDisposal.asset_id == FixedAsset.id)
            .where(
                and_(
                    AssetDisposal.entity_id == entity_id,
                    extract('year', AssetDisposal.disposal_date) == year
                )
            )
            .order_by(AssetDisposal.disposal_date)
        )
        disposals = result.all()
        
        row = 2
        total_cost = Decimal("0")
        total_accum_dep = Decimal("0")
        total_nbv = Decimal("0")
        total_proceeds = Decimal("0")
        total_gain_loss = Decimal("0")
        
        for disposal, asset in disposals:
            ws.cell(row=row, column=1).value = asset.asset_number
            ws.cell(row=row, column=2).value = asset.asset_name
            ws.cell(row=row, column=3).value = disposal.disposal_date
            ws.cell(row=row, column=3).number_format = 'MM/DD/YYYY'
            ws.cell(row=row, column=4).value = disposal.disposal_type
            ws.cell(row=row, column=5).value = float(disposal.original_cost)
            ws.cell(row=row, column=5).number_format = '$#,##0.00'
            ws.cell(row=row, column=6).value = float(disposal.accumulated_depreciation)
            ws.cell(row=row, column=6).number_format = '$(#,##0.00)'
            ws.cell(row=row, column=7).value = float(disposal.net_book_value)
            ws.cell(row=row, column=7).number_format = '$#,##0.00'
            ws.cell(row=row, column=8).value = float(disposal.disposal_amount)
            ws.cell(row=row, column=8).number_format = '$#,##0.00'
            ws.cell(row=row, column=9).value = float(disposal.gain_loss)
            ws.cell(row=row, column=9).number_format = '$#,##0.00;[Red]$(#,##0.00)'
            ws.cell(row=row, column=10).value = disposal.disposal_notes or ""
            
            total_cost += disposal.original_cost
            total_accum_dep += disposal.accumulated_depreciation
            total_nbv += disposal.net_book_value
            total_proceeds += disposal.disposal_amount
            total_gain_loss += disposal.gain_loss
            
            row += 1
        
        # Totals
        if row > 2:
            ws.cell(row=row, column=1).value = "TOTAL"
            ws.cell(row=row, column=1).font = AuditPackageGenerator.TOTAL_FONT
            ws.cell(row=row, column=5).value = float(total_cost)
            ws.cell(row=row, column=5).number_format = '$#,##0.00'
            ws.cell(row=row, column=5).font = AuditPackageGenerator.TOTAL_FONT
            ws.cell(row=row, column=6).value = float(total_accum_dep)
            ws.cell(row=row, column=6).number_format = '$(#,##0.00)'
            ws.cell(row=row, column=6).font = AuditPackageGenerator.TOTAL_FONT
            ws.cell(row=row, column=7).value = float(total_nbv)
            ws.cell(row=row, column=7).number_format = '$#,##0.00'
            ws.cell(row=row, column=7).font = AuditPackageGenerator.TOTAL_FONT
            ws.cell(row=row, column=8).value = float(total_proceeds)
            ws.cell(row=row, column=8).number_format = '$#,##0.00'
            ws.cell(row=row, column=8).font = AuditPackageGenerator.TOTAL_FONT
            ws.cell(row=row, column=9).value = float(total_gain_loss)
            ws.cell(row=row, column=9).number_format = '$#,##0.00;[Red]$(#,##0.00)'
            ws.cell(row=row, column=9).font = AuditPackageGenerator.TOTAL_FONT
        else:
            ws.cell(row=row, column=1).value = "No disposals in this period"
        
        # Format columns
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 18
        ws.column_dimensions['G'].width = 15
        ws.column_dimensions['H'].width = 15
        ws.column_dimensions['I'].width = 15
        ws.column_dimensions['J'].width = 35
        
        ws.freeze_panes = 'A2'


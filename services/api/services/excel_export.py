"""
NGI Capital - Excel Export Service
Generates investor-ready financial statement packages in Excel format

Implements Deloitte EGC (Emerging Growth Company) template format

Author: NGI Capital Development Team
Date: October 3, 2025
"""

from datetime import date
from decimal import Decimal
from typing import Dict, Any
import io
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill, NamedStyle
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    Workbook = None


class ExcelFinancialStatementExporter:
    """
    Exports financial statements to Excel using Deloitte EGC format
    Professional formatting for investor presentations
    """
    
    def __init__(self, financial_data: Dict[str, Any]):
        if not OPENPYXL_AVAILABLE:
            raise ImportError("openpyxl is required for Excel export")
        
        self.data = financial_data
        self.wb = Workbook()
        
        # Define styles
        self._setup_styles()
    
    def _setup_styles(self):
        """Setup professional Excel styles matching Deloitte EGC format"""
        
        # Header style (Company name)
        self.header_style = NamedStyle(name="header")
        self.header_style.font = Font(name="Arial", size=14, bold=True)
        self.header_style.alignment = Alignment(horizontal="center", vertical="center")
        
        # Statement title style
        self.title_style = NamedStyle(name="title")
        self.title_style.font = Font(name="Arial", size=12, bold=True)
        self.title_style.alignment = Alignment(horizontal="center", vertical="center")
        
        # Section header style
        self.section_header_style = NamedStyle(name="section_header")
        self.section_header_style.font = Font(name="Arial", size=10, bold=True)
        self.section_header_style.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        
        # Account line style
        self.account_style = NamedStyle(name="account")
        self.account_style.font = Font(name="Arial", size=10)
        self.account_style.alignment = Alignment(horizontal="left", indent=1)
        
        # Number style
        self.number_style = NamedStyle(name="number")
        self.number_style.font = Font(name="Arial", size=10)
        self.number_style.alignment = Alignment(horizontal="right")
        self.number_style.number_format = "#,##0.00"
        
        # Total line style
        self.total_style = NamedStyle(name="total")
        self.total_style.font = Font(name="Arial", size=10, bold=True)
        self.total_style.fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
        self.total_style.border = Border(
            top=Side(style="thin"),
            bottom=Side(style="double", border_style="double")
        )
        
        # Add styles to workbook
        try:
            self.wb.add_named_style(self.header_style)
            self.wb.add_named_style(self.title_style)
            self.wb.add_named_style(self.section_header_style)
            self.wb.add_named_style(self.account_style)
            self.wb.add_named_style(self.number_style)
            self.wb.add_named_style(self.total_style)
        except:
            pass  # Styles may already exist
    
    def generate_workbook(self) -> io.BytesIO:
        """Generate complete Excel workbook with all statements"""
        
        # Remove default sheet
        if "Sheet" in self.wb.sheetnames:
            self.wb.remove(self.wb["Sheet"])
        
        # Create sheets for each statement
        self._create_cover_sheet()
        self._create_balance_sheet()
        self._create_income_statement()
        self._create_cash_flows()
        self._create_equity_statement()
        self._create_comprehensive_income()
        self._create_notes()
        
        # Save to BytesIO
        output = io.BytesIO()
        self.wb.save(output)
        output.seek(0)
        
        return output
    
    def _create_cover_sheet(self):
        """Create cover sheet with company info"""
        
        ws = self.wb.create_sheet("Cover", 0)
        
        # Company name
        ws["B2"] = self.data["entity_name"]
        ws["B2"].style = "header"
        ws.merge_cells("B2:E2")
        
        # Title
        ws["B4"] = "Financial Statements"
        ws["B4"].style = "title"
        ws.merge_cells("B4:E4")
        
        # Period
        ws["B6"] = f"For the period ending {self.data['period_end_date']}"
        ws["B6"].font = Font(name="Arial", size=11)
        ws["B6"].alignment = Alignment(horizontal="center")
        ws.merge_cells("B6:E6")
        
        # Contents
        ws["B9"] = "Table of Contents"
        ws["B9"].font = Font(name="Arial", size=12, bold=True)
        
        contents = [
            "Balance Sheet",
            "Income Statement",
            "Cash Flows",
            "Stockholders' Equity",
            "Comprehensive Income",
            "Notes to Financial Statements"
        ]
        
        for idx, item in enumerate(contents, start=11):
            ws[f"B{idx}"] = f"{idx-10}. {item}"
            ws[f"B{idx}"].font = Font(name="Arial", size=10)
        
        # Set column widths
        ws.column_dimensions["B"].width = 50
    
    def _create_balance_sheet(self):
        """Create balance sheet"""
        
        ws = self.wb.create_sheet("Balance Sheet")
        bs = self.data["statements"]["balance_sheet"]
        
        row = 1
        
        # Header
        ws[f"A{row}"] = self.data["entity_name"]
        ws[f"A{row}"].style = "header"
        ws.merge_cells(f"A{row}:C{row}")
        row += 1
        
        ws[f"A{row}"] = "Balance Sheet"
        ws[f"A{row}"].style = "title"
        ws.merge_cells(f"A{row}:C{row}")
        row += 1
        
        ws[f"A{row}"] = f"As of {self.data['period_end_date']}"
        ws[f"A{row}"].alignment = Alignment(horizontal="center")
        ws.merge_cells(f"A{row}:C{row}")
        row += 2
        
        # Assets
        ws[f"A{row}"] = "ASSETS"
        ws[f"A{row}"].style = "section_header"
        ws.merge_cells(f"A{row}:C{row}")
        row += 1
        
        # Current Assets
        ws[f"A{row}"] = "Current Assets"
        ws[f"A{row}"].font = Font(bold=True)
        row += 1
        
        for account_name, details in bs["assets"]["current_assets"].items():
            ws[f"A{row}"] = f"  {account_name}"
            ws[f"C{row}"] = details["balance"]
            ws[f"C{row}"].style = "number"
            row += 1
        
        ws[f"A{row}"] = "Total Current Assets"
        ws[f"A{row}"].font = Font(bold=True, italic=True)
        row += 1
        
        # Non-current Assets
        if bs["assets"]["non_current_assets"]:
            ws[f"A{row}"] = "Non-Current Assets"
            ws[f"A{row}"].font = Font(bold=True)
            row += 1
            
            for account_name, details in bs["assets"]["non_current_assets"].items():
                ws[f"A{row}"] = f"  {account_name}"
                ws[f"C{row}"] = details["balance"]
                ws[f"C{row}"].style = "number"
                row += 1
        
        # Total Assets
        ws[f"A{row}"] = "TOTAL ASSETS"
        ws[f"C{row}"] = bs["assets"]["total_assets"]
        ws[f"A{row}"].style = "total"
        ws[f"C{row}"].style = "total"
        row += 2
        
        # Liabilities
        ws[f"A{row}"] = "LIABILITIES AND STOCKHOLDERS' EQUITY"
        ws[f"A{row}"].style = "section_header"
        ws.merge_cells(f"A{row}:C{row}")
        row += 1
        
        # Current Liabilities
        ws[f"A{row}"] = "Current Liabilities"
        ws[f"A{row}"].font = Font(bold=True)
        row += 1
        
        for account_name, details in bs["liabilities"]["current_liabilities"].items():
            ws[f"A{row}"] = f"  {account_name}"
            ws[f"C{row}"] = details["balance"]
            ws[f"C{row}"].style = "number"
            row += 1
        
        # Total Liabilities
        ws[f"A{row}"] = "Total Liabilities"
        ws[f"C{row}"] = bs["liabilities"]["total_liabilities"]
        ws[f"A{row}"].font = Font(bold=True)
        ws[f"C{row}"].style = "number"
        row += 2
        
        # Stockholders' Equity
        ws[f"A{row}"] = "Stockholders' Equity"
        ws[f"A{row}"].font = Font(bold=True)
        row += 1
        
        for account_name, details in bs["equity"]["stockholders_equity"].items():
            ws[f"A{row}"] = f"  {account_name}"
            ws[f"C{row}"] = details["balance"]
            ws[f"C{row}"].style = "number"
            row += 1
        
        ws[f"A{row}"] = "Total Stockholders' Equity"
        ws[f"C{row}"] = bs["equity"]["total_equity"]
        ws[f"A{row}"].font = Font(bold=True)
        ws[f"C{row}"].style = "number"
        row += 2
        
        # Total Liabilities and Equity
        total_liab_equity = bs["liabilities"]["total_liabilities"] + bs["equity"]["total_equity"]
        ws[f"A{row}"] = "TOTAL LIABILITIES AND STOCKHOLDERS' EQUITY"
        ws[f"C{row}"] = total_liab_equity
        ws[f"A{row}"].style = "total"
        ws[f"C{row}"].style = "total"
        
        # Set column widths
        ws.column_dimensions["A"].width = 50
        ws.column_dimensions["B"].width = 5
        ws.column_dimensions["C"].width = 20
    
    def _create_income_statement(self):
        """Create income statement with expense disaggregation"""
        
        ws = self.wb.create_sheet("Income Statement")
        inc = self.data["statements"]["income_statement"]
        
        row = 1
        
        # Header
        ws[f"A{row}"] = self.data["entity_name"]
        ws[f"A{row}"].style = "header"
        ws.merge_cells(f"A{row}:C{row}")
        row += 1
        
        ws[f"A{row}"] = "Statement of Operations"
        ws[f"A{row}"].style = "title"
        ws.merge_cells(f"A{row}:C{row}")
        row += 1
        
        ws[f"A{row}"] = f"For the period from {self.data['period_start_date']} to {self.data['period_end_date']}"
        ws[f"A{row}"].alignment = Alignment(horizontal="center")
        ws.merge_cells(f"A{row}:C{row}")
        row += 2
        
        # Revenue
        ws[f"A{row}"] = "REVENUE"
        ws[f"A{row}"].style = "section_header"
        row += 1
        
        for account_name, details in inc["revenue"].items():
            ws[f"A{row}"] = f"  {account_name}"
            ws[f"C{row}"] = details["amount"]
            ws[f"C{row}"].style = "number"
            row += 1
        
        # Cost of Revenue
        ws[f"A{row}"] = "COST OF REVENUE"
        ws[f"A{row}"].font = Font(bold=True)
        row += 1
        
        for account_name, details in inc["cost_of_revenue"].items():
            ws[f"A{row}"] = f"  {account_name}"
            ws[f"C{row}"] = details["amount"]
            ws[f"C{row}"].style = "number"
            row += 1
        
        # Gross Profit
        ws[f"A{row}"] = "Gross Profit"
        ws[f"C{row}"] = inc["gross_profit"]
        ws[f"A{row}"].style = "total"
        ws[f"C{row}"].style = "total"
        row += 2
        
        # Operating Expenses (Disaggregated per 2025 GAAP)
        ws[f"A{row}"] = "OPERATING EXPENSES"
        ws[f"A{row}"].style = "section_header"
        row += 1
        
        # R&D
        if inc["operating_expenses"]["research_and_development"]:
            ws[f"A{row}"] = "Research and Development"
            ws[f"A{row}"].font = Font(bold=True)
            row += 1
            for account_name, details in inc["operating_expenses"]["research_and_development"].items():
                ws[f"A{row}"] = f"  {account_name}"
                ws[f"C{row}"] = details["amount"]
                ws[f"C{row}"].style = "number"
                row += 1
        
        # Sales & Marketing
        if inc["operating_expenses"]["sales_and_marketing"]:
            ws[f"A{row}"] = "Sales and Marketing"
            ws[f"A{row}"].font = Font(bold=True)
            row += 1
            for account_name, details in inc["operating_expenses"]["sales_and_marketing"].items():
                ws[f"A{row}"] = f"  {account_name}"
                ws[f"C{row}"] = details["amount"]
                ws[f"C{row}"].style = "number"
                row += 1
        
        # G&A
        if inc["operating_expenses"]["general_and_administrative"]:
            ws[f"A{row}"] = "General and Administrative"
            ws[f"A{row}"].font = Font(bold=True)
            row += 1
            for account_name, details in inc["operating_expenses"]["general_and_administrative"].items():
                ws[f"A{row}"] = f"  {account_name}"
                ws[f"C{row}"] = details["amount"]
                ws[f"C{row}"].style = "number"
                row += 1
        
        # Total Operating Expenses
        ws[f"A{row}"] = "Total Operating Expenses"
        ws[f"C{row}"] = inc["operating_expenses"]["total_operating_expenses"]
        ws[f"A{row}"].font = Font(bold=True)
        ws[f"C{row}"].style = "number"
        row += 2
        
        # Operating Income
        ws[f"A{row}"] = "Operating Income"
        ws[f"C{row}"] = inc["operating_income"]
        ws[f"A{row}"].style = "total"
        ws[f"C{row}"].style = "total"
        row += 2
        
        # Net Income
        ws[f"A{row}"] = "NET INCOME"
        ws[f"C{row}"] = inc["net_income"]
        ws[f"A{row}"].style = "total"
        ws[f"C{row}"].style = "total"
        
        # Set column widths
        ws.column_dimensions["A"].width = 50
        ws.column_dimensions["B"].width = 5
        ws.column_dimensions["C"].width = 20
    
    def _create_cash_flows(self):
        """Create statement of cash flows"""
        
        ws = self.wb.create_sheet("Cash Flows")
        cf = self.data["statements"]["cash_flows"]
        
        row = 1
        
        # Header
        ws[f"A{row}"] = self.data["entity_name"]
        ws[f"A{row}"].style = "header"
        ws.merge_cells(f"A{row}:C{row}")
        row += 1
        
        ws[f"A{row}"] = "Statement of Cash Flows"
        ws[f"A{row}"].style = "title"
        ws.merge_cells(f"A{row}:C{row}")
        row += 1
        
        ws[f"A{row}"] = f"For the period from {self.data['period_start_date']} to {self.data['period_end_date']}"
        ws[f"A{row}"].alignment = Alignment(horizontal="center")
        ws.merge_cells(f"A{row}:C{row}")
        row += 2
        
        # Operating Activities
        ws[f"A{row}"] = "CASH FLOWS FROM OPERATING ACTIVITIES"
        ws[f"A{row}"].style = "section_header"
        row += 1
        
        ws[f"A{row}"] = "  Net Income"
        ws[f"C{row}"] = cf["operating_activities"]["net_income"]
        ws[f"C{row}"].style = "number"
        row += 2
        
        ws[f"A{row}"] = "Net Cash from Operating Activities"
        ws[f"C{row}"] = cf["operating_activities"]["net_cash_from_operating"]
        ws[f"A{row}"].font = Font(bold=True)
        ws[f"C{row}"].style = "number"
        row += 2
        
        # Net Change in Cash
        ws[f"A{row}"] = "NET CHANGE IN CASH"
        ws[f"C{row}"] = cf["net_change_in_cash"]
        ws[f"A{row}"].style = "total"
        ws[f"C{row}"].style = "total"
        row += 2
        
        # Beginning and Ending Cash
        ws[f"A{row}"] = "Cash, beginning of period"
        ws[f"C{row}"] = cf["cash_beginning"]
        ws[f"C{row}"].style = "number"
        row += 1
        
        ws[f"A{row}"] = "Cash, end of period"
        ws[f"C{row}"] = cf["cash_ending"]
        ws[f"A{row}"].style = "total"
        ws[f"C{row}"].style = "total"
        
        # Set column widths
        ws.column_dimensions["A"].width = 50
        ws.column_dimensions["B"].width = 5
        ws.column_dimensions["C"].width = 20
    
    def _create_equity_statement(self):
        """Create statement of stockholders' equity"""
        
        ws = self.wb.create_sheet("Stockholders Equity")
        
        # Placeholder - simplified version
        ws["A1"] = self.data["entity_name"]
        ws["A1"].style = "header"
        ws["A3"] = "Statement of Stockholders' Equity"
        ws["A3"].style = "title"
        ws["A5"] = "To be implemented"
    
    def _create_comprehensive_income(self):
        """Create statement of comprehensive income"""
        
        ws = self.wb.create_sheet("Comprehensive Income")
        
        # Placeholder - simplified version
        ws["A1"] = self.data["entity_name"]
        ws["A1"].style = "header"
        ws["A3"] = "Statement of Comprehensive Income"
        ws["A3"].style = "title"
        ws["A5"] = "To be implemented"
    
    def _create_notes(self):
        """Create notes to financial statements"""
        
        ws = self.wb.create_sheet("Notes")
        
        row = 1
        
        # Header
        ws[f"A{row}"] = self.data["entity_name"]
        ws[f"A{row}"].style = "header"
        ws.merge_cells(f"A{row}:D{row}")
        row += 1
        
        ws[f"A{row}"] = "Notes to Financial Statements"
        ws[f"A{row}"].style = "title"
        ws.merge_cells(f"A{row}:D{row}")
        row += 2
        
        # Notes
        for note in self.data["notes"]:
            ws[f"A{row}"] = f"Note {note['note_number']}: {note['title']}"
            ws[f"A{row}"].font = Font(name="Arial", size=11, bold=True)
            row += 1
            
            ws[f"A{row}"] = note["content"]
            ws[f"A{row}"].font = Font(name="Arial", size=10)
            ws[f"A{row}"].alignment = Alignment(wrap_text=True)
            ws.merge_cells(f"A{row}:D{row}")
            row += 2
        
        # Set column widths
        ws.column_dimensions["A"].width = 80

